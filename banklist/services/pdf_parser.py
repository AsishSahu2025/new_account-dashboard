"""
services/pdf_parser.py
======================
Bank Statement PDF Parser — tested against real DBS, SBI, HDFC statements.

Strategy per bank
-----------------
DBS   → word-position parser  (columns identified by x0 coordinate)
HDFC  → word-position parser  (columns identified by x0 coordinate)
SBI   → pdfplumber table extraction (pdfplumber detects the table perfectly)
ICICI → table extraction fallback
AXIS  → table extraction fallback
Generic → table → text fallback chain

All parsers return the same normalised ParsedTransaction dataclass so the
rest of your system (views.py, reconciliation engine) never needs to change
when a new bank is added.
"""

import io
import re
import logging
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import List, Optional

import pdfplumber

from .ollama_extractor import OllamaExtractionError, extract_statement_transactions

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Data classes
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ParsedTransaction:
    txn_date:   date
    value_date: Optional[date]
    description: str
    ref_no:     str
    utr_no:     str
    txn_type:   str          # 'debit' | 'credit'
    amount:     Decimal
    balance:    Optional[Decimal]
    cheque_no:  str = ''


@dataclass
class ParseResult:
    bank_name:        str
    account_no:       str
    account_name:     str
    transactions:     List[ParsedTransaction] = field(default_factory=list)
    errors:           List[str]               = field(default_factory=list)
    statement_period: dict                    = field(default_factory=dict)


# ─────────────────────────────────────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────────────────────────────────────

def _parse_amount(val) -> Optional[Decimal]:
    """'1,23,456.78' / '-29,51,582.88' / '₹1000.00' → Decimal or None."""
    if not val:
        return None
    cleaned = re.sub(r'[₹$,\s]', '', str(val).strip())
    # keep a leading minus (negative balances)
    try:
        d = Decimal(cleaned)
        return d
    except (InvalidOperation, ValueError):
        return None


def _parse_amount_positive(val) -> Optional[Decimal]:
    """Same as above but only returns value if > 0 (used for debit/credit)."""
    d = _parse_amount(val)
    return d if d and d > 0 else None


def _parse_date(s, fmts: List[str]) -> Optional[date]:
    s = str(s).strip() if s else ''
    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _extract_utr(text: str) -> str:
    """Pull UTR / NEFT reference from description text."""
    patterns = [
        r'\b(SBIN\d{9,18})\b',
        r'\b(HDFC[A-Z0-9]{6,18})\b',
        r'\b(DBSS[A-Z0-9]{6,15})\b',
        r'\b(UTIB[A-Z0-9]{6,15})\b',
        r'\b(CNAE[A-Z0-9]{4,12})\b',
        r'UTR\s*(?:NO)?[:\s]*([A-Z0-9]{8,25})',
        r'NEFT\s+(?:UTR\s+)?NO[:\s]*([A-Z0-9]{8,25})',
        r'\b([A-Z]{4}[A-Z0-9]{6,20})\b',
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return m.group(1).upper()
    return ''


def _clean_description(text: str) -> str:
    return ' '.join(text.split())[:500]


def _deduplicate(txns: List[ParsedTransaction]) -> List[ParsedTransaction]:
    """
    De-dupe parsed rows.

    IMPORTANT: the key is anchored on the running `balance`, not just a
    truncated description. On real statements it's common to have two
    genuinely different transactions on the same date, for the same
    amount, to/from the same counterparty (e.g. two separate UPI debits
    to the same person for different purposes) — their first ~30 chars
    of description are identical, so keying on description[:30] alone
    silently drops one of them. The running balance changes with every
    transaction in a real ledger, so it's a much safer differentiator.
    ref_no (when present) is the next-best differentiator; description
    is only used as a last-resort fallback when neither is available.
    """
    seen, out = set(), []
    for t in txns:
        key = (
            t.txn_date.isoformat(),
            str(t.amount),
            t.txn_type,
            str(t.balance) if t.balance is not None else '',
            t.ref_no or t.description[:50],
        )
        if key not in seen:
            seen.add(key)
            out.append(t)
    return out


def _parse_iso_date(value: Optional[str]) -> Optional[date]:
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


def _extract_pdf_text_with_ocr(file_bytes: bytes) -> str:
    """Fallback text extraction for scanned statement PDFs."""
    try:
        import fitz
        import pytesseract
        from PIL import Image

        doc = fitz.open(stream=file_bytes, filetype="pdf")
        pages = []
        for page in doc:
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            img = Image.open(io.BytesIO(pix.tobytes("png")))
            pages.append(pytesseract.image_to_string(img, lang="eng"))
        return "\n".join(pages).strip()
    except Exception as exc:
        logger.warning("Statement OCR fallback failed: %s", exc)
        return ""


def _is_excel_file(file_name: str) -> bool:
    return str(file_name or '').lower().endswith(('.xlsx', '.xlsm'))


def _is_image_file(file_name: str) -> bool:
    return str(file_name or '').lower().endswith(('.png', '.jpg', '.jpeg'))


def _validate_pdf_file(file_bytes: bytes) -> tuple:
    """
    Validate that the uploaded file is a readable PDF.

    Returns:
        (True, None) when valid.
        (False, error_message) when invalid.
    """

    if not file_bytes:
        return False, 'No file was provided.'

    # Check PDF file signature
    if not file_bytes.startswith(b'%PDF-'):
        return False, 'The uploaded file is not a valid PDF.'

    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            if not pdf.pages:
                return False, 'The uploaded PDF contains no pages.'

    except Exception:
        logger.warning('Unable to open uploaded PDF', exc_info=True)
        return False, 'The uploaded PDF is corrupted or cannot be opened.'

    return True, None


def _extract_excel_text(file_bytes: bytes) -> str:
    """Extract workbook cell text for AI-assisted statement parsing."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        chunks = []
        for sheet in wb.worksheets:
            chunks.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [str(cell).strip() for cell in row if cell not in (None, '')]
                if values:
                    chunks.append(" | ".join(values))
        return "\n".join(chunks).strip()
    except Exception as exc:
        logger.warning("Excel statement extraction failed: %s", exc)
        return ""


def _extract_image_text(file_bytes: bytes) -> str:
    """OCR an image bank statement so Ollama can extract transaction rows."""
    try:
        import pytesseract
        from PIL import Image, ImageEnhance

        img = Image.open(io.BytesIO(file_bytes))
        if img.mode != "L":
            img = img.convert("L")
        img = ImageEnhance.Contrast(img).enhance(2.0)
        return pytesseract.image_to_string(img, lang="eng").strip()
    except Exception as exc:
        logger.warning("Image statement OCR failed: %s", exc)
        return ""


def _ollama_statement_result(text: str) -> Optional[ParseResult]:
    if not text.strip():
        return None
    try:
        ai = extract_statement_transactions(text)
    except OllamaExtractionError as exc:
        logger.warning("Ollama statement fallback skipped: %s", exc)
        return None
    except Exception as exc:
        logger.warning("Ollama statement fallback failed: %s", exc, exc_info=True)
        return None

    txns = []
    for item in ai.get("transactions") or []:
        txn_date = _parse_iso_date(item.get("txn_date"))
        if not txn_date:
            continue
        amount = item.get("amount")
        if not amount or amount <= 0:
            continue
        txns.append(ParsedTransaction(
            txn_date=txn_date,
            value_date=_parse_iso_date(item.get("value_date")),
            description=_clean_description(item.get("description") or ""),
            ref_no=item.get("ref_no") or "",
            utr_no=(item.get("utr_no") or _extract_utr(item.get("description") or "")),
            txn_type=item.get("txn_type"),
            amount=amount,
            balance=item.get("balance"),
        ))

    if not txns:
        return None

    return ParseResult(
        bank_name=ai.get("bank_name") or "UNKNOWN",
        account_no=ai.get("account_no") or "",
        account_name=ai.get("account_name") or "",
        transactions=_deduplicate(txns),
        errors=[],
    )


# ─────────────────────────────────────────────────────────────────────────────
# Bank detection + metadata
# ─────────────────────────────────────────────────────────────────────────────

_BANK_SIGNATURES = {
    # !! ORDER MATTERS !!
    # DBS   — unique: "EBUSINESS LITE" product type, 8160210 account prefix
    # HDFC  — unique: "WithdrawalAmt" / "DepositAmt" (words run together in their PDF)
    #          Must come BEFORE SBI because SBI narrations mention HDFC destinations
    # SBI   — unique: "IFS Code" header field contains SBIN
    # ICICI / AXIS — generic fallback
    'DBS':   [r'EBUSINESS LITE', r'DBS BANK', r'8160210', r'DBSS\d{4}'],
    'HDFC':  [r'WithdrawalAmt', r'DepositAmt', r'ASCENTCURRENT', r'HDFC0004\d+'],
    'SBI':   [r'STATE BANK OF INDIA', r'IFS\s*Code\s*[:\s]+SBIN', r'SBIN0010\d+'],
    'ICICI': [r'ICICI BANK', r'ICIC\d{4}'],
    'AXIS':  [r'AXIS BANK', r'UTIB\d{4}'],
}


def _detect_bank(text: str) -> str:
    upper = text.upper()
    for bank, patterns in _BANK_SIGNATURES.items():
        if any(re.search(p, upper) for p in patterns):
            return bank
    return 'UNKNOWN'


def _extract_metadata(text: str) -> tuple:
    """Return (account_no, account_name)."""
    acc_no = ''
    acc_name = ''

    for pat in [
        r'Account\s*Number\s*[:\-]?\s*([\d\-]{8,25})',
        r'A/C\s*No\.?\s*[:\-]?\s*([\d\-]{8,25})',
        r'Account\s*No\.?\s*[:\-]?\s*([\d\-]{8,25})',
        r'(\d{10,20})',
    ]:
        m = re.search(pat, text, re.I)
        if m:
            acc_no = m.group(1).strip()
            break

    for pat in [
        r'Account\s*Name\s*[:\-]?\s*([^\n\r:]+)',
        r'Customer\s*Name\s*[:\-]?\s*([^\n\r:]+)',
        r'M/S\.?\s+([^\n\r]+)',
    ]:
        m = re.search(pat, text, re.I)
        if m:
            acc_name = m.group(1).strip()
            break

    return acc_no, acc_name


# ─────────────────────────────────────────────────────────────────────────────
# DBS parser  (word-position based)
#
# Column boundaries re-measured directly against a real DBS "Account
# Details" export (595pt-wide A4 page). The previous boundaries were a
# rough guess and didn't match this layout at all — e.g. the old
# 'details' band (150, 310) missed the real Transaction Details column
# (x0≈195), and the old 'debits'/'credits' bands (310-430 / 430-520)
# didn't line up with where amounts are actually right-aligned
# (Debit x0≈369-390, x1≈405 | Credit x0≈420-450, x1≈468).
#
# New boundaries (verified with pdfplumber word coordinates,
# x_tolerance=3, against every row of a real statement):
#   Date               x0 ≈ 37             → band (0, 100)
#   Value Date         x0 ≈ 116-155        → band (100, 190)
#   Transaction Details x0 ≈ 195, wraps to x1≈340 → band (190, 355)
#   Debit              x0 ≈ 369-390, x1≈405        → band (355, 420)
#   Credit             x0 ≈ 420-450, x1≈468         → band (420, 490)
#   Running Balance    x0 ≈ 522-534, x1≈558         → band (490, 600)
# ─────────────────────────────────────────────────────────────────────────────

_DBS_COLS = {
    'trans_date':  (0,   100),
    'value_date':  (100, 190),
    'details':     (190, 355),
    'debits':      (355, 420),
    'credits':     (420, 490),
    'balance':     (490, 600),
}
_DBS_DATE_FMTS = ['%d-%b-%Y', '%d/%m/%Y', '%d-%m-%Y']


def _in_dbs(word: dict, col: str) -> bool:
    x0, x1 = _DBS_COLS[col]
    return x0 <= word['x0'] < x1


def _parse_dbs(pdf) -> tuple:
    txns, errors = [], []

    for page_num, page in enumerate(pdf.pages, 1):
        try:
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            rows: dict = defaultdict(list)
            for w in words:
                rows[round(w['top'])].append(w)

            pending = None

            for top in sorted(rows):
                rw = sorted(rows[top], key=lambda w: w['x0'])
                if not rw:
                    continue

                first = rw[0]

                # Sentinel: the transaction table ends at the "Total Debit
                # Count / Total Credit Count / Total ... Amount" summary
                # rows. Everything below that (disclaimers, "**END OF
                # REPORT**", "Printed By ... Page 1 / 1") is footer text —
                # without this check, footer fragments like the page
                # number "1" in "Page 1 / 1" can land inside the credit
                # column's x-range and get misread as a stray credit
                # amount tacked onto the last real transaction.
                if first['text'] == 'Total':
                    break

                txn_date = (_in_dbs(first, 'trans_date') and
                            _parse_date(first['text'], _DBS_DATE_FMTS))

                if txn_date:
                    # Save previous
                    if pending:
                        t = _build_dbs_txn(pending)
                        if t:
                            txns.append(t)
                    # New transaction
                    pending = {
                        'date': txn_date,
                        'desc': [],
                        'debit': None,
                        'credit': None,
                        'balance': None,
                    }
                    for w in rw[1:]:
                        txt = w['text']
                        if _in_dbs(w, 'details'):
                            pending['desc'].append(txt)
                        elif _in_dbs(w, 'debits'):
                            a = _parse_amount_positive(txt)
                            if a:
                                pending['debit'] = a
                        elif _in_dbs(w, 'credits'):
                            a = _parse_amount_positive(txt)
                            if a:
                                pending['credit'] = a
                        elif _in_dbs(w, 'balance'):
                            # balance can be negative
                            a = _parse_amount(txt.replace('-', ''))
                            if a:
                                pending['balance'] = a
                else:
                    # Continuation row
                    if pending:
                        for w in rw:
                            txt = w['text']
                            if _in_dbs(w, 'details'):
                                pending['desc'].append(txt)
                            elif _in_dbs(w, 'debits') and pending['debit'] is None:
                                a = _parse_amount_positive(txt)
                                if a:
                                    pending['debit'] = a
                            elif _in_dbs(w, 'credits') and pending['credit'] is None:
                                a = _parse_amount_positive(txt)
                                if a:
                                    pending['credit'] = a
                            elif _in_dbs(w, 'balance') and pending['balance'] is None:
                                a = _parse_amount(txt.replace('-', ''))
                                if a:
                                    pending['balance'] = a

            if pending:
                t = _build_dbs_txn(pending)
                if t:
                    txns.append(t)

        except Exception as e:
            errors.append(f'DBS page {page_num}: {e}')
            logger.warning('DBS page %d error: %s', page_num, e)

    return txns, errors


def _build_dbs_txn(p: dict) -> Optional[ParsedTransaction]:
    if not p.get('date'):
        return None
    debit  = p.get('debit')
    credit = p.get('credit')

    if debit and debit > 0:
        txn_type, amount = 'debit', debit
    elif credit and credit > 0:
        txn_type, amount = 'credit', credit
    else:
        return None

    desc = _clean_description(' '.join(p['desc']))
    utr  = _extract_utr(desc)

    return ParsedTransaction(
        txn_date=p['date'],
        value_date=None,
        description=desc,
        ref_no=utr,
        utr_no=utr,
        txn_type=txn_type,
        amount=amount,
        balance=p.get('balance'),
    )


# ─────────────────────────────────────────────────────────────────────────────
# SBI parser  (pdfplumber table extraction — works perfectly on SBI PDFs)
# Columns: Txn Date | Value Date | Description | Ref No./Cheque No. |
#          Branch Code | Debit | Credit | Balance
# ─────────────────────────────────────────────────────────────────────────────

_SBI_DATE_FMTS = ['%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y']

# keywords → standard field name
_SBI_HEADER_MAP = {
    'txn date': 'date',
    # 'date': 'date',
    'value date': 'vdate',
    'description': 'desc',
    'ref no': 'ref',
    'cheque': 'ref',
    'debit': 'debit',
    'credit': 'credit',
    'balance': 'balance',
}


def _map_sbi_headers(headers: list) -> dict:
    """Return {field: col_index} from header list."""
    col = {}
    for i, h in enumerate(headers):
        h_clean = h.lower().replace('\n', ' ').strip()
        for keyword, field in _SBI_HEADER_MAP.items():
            if keyword in h_clean:
                if field not in col:          # first match wins
                    col[field] = i
                break
    return col


def _parse_sbi(pdf) -> tuple:
    txns, errors = [], []

    for page_num, page in enumerate(pdf.pages, 1):
        try:
            tables = page.extract_tables()
            for table in tables:
                if not table or len(table) < 2:
                    continue

                # Find header row
                hdr_idx = None
                for i, row in enumerate(table):
                    joined = ' '.join(str(c) for c in row if c).upper()
                    if ('DEBIT' in joined and 'CREDIT' in joined) or 'TXN DATE' in joined:
                        hdr_idx = i
                        break
                if hdr_idx is None:
                    continue

                headers = [str(c or '').strip() for c in table[hdr_idx]]
                col = _map_sbi_headers(headers)

                if 'date' not in col:
                    continue

                for row in table[hdr_idx + 1:]:
                    if not row:
                        continue
                    try:
                        date_raw = str(row[col['date']] or '').split('\n')[0].strip()
                        txn_date = _parse_date(date_raw, _SBI_DATE_FMTS)
                        if not txn_date:
                            continue

                        vdate = None
                        if 'vdate' in col:
                            vdate = _parse_date(
                                str(row[col['vdate']] or '').split('\n')[0].strip(),
                                _SBI_DATE_FMTS
                            )

                        desc = _clean_description(
                            str(row[col.get('desc', 2)] or '').replace('\n', ' ')
                        )
                        ref_raw = str(row[col.get('ref', 3)] or '').replace('\n', ' ').strip()
                        # First token of ref column is usually the UTR/ref
                        ref_no = ref_raw.split()[0] if ref_raw else ''

                        debit  = _parse_amount_positive(row[col['debit']])  if 'debit'  in col else None
                        credit = _parse_amount_positive(row[col['credit']]) if 'credit' in col else None

                        # Balance may be negative (credit card / OD account)
                        bal = None
                        if 'balance' in col:
                            bal = _parse_amount(str(row[col['balance']] or '').replace(',', '').replace(' ', ''))

                        if debit and debit > 0:
                            txn_type, amount = 'debit', debit
                        elif credit and credit > 0:
                            txn_type, amount = 'credit', credit
                        else:
                            continue

                        utr = _extract_utr(desc + ' ' + ref_no)

                        txns.append(ParsedTransaction(
                            txn_date=txn_date,
                            value_date=vdate,
                            description=desc,
                            ref_no=ref_no,
                            utr_no=utr,
                            txn_type=txn_type,
                            amount=amount,
                            balance=bal,
                        ))
                    except Exception as row_err:
                        logger.debug('SBI row error: %s', row_err)

        except Exception as e:
            errors.append(f'SBI page {page_num}: {e}')
            logger.warning('SBI page %d error: %s', page_num, e)

    return txns, errors


# ─────────────────────────────────────────────────────────────────────────────
# HDFC parser  (word-position based)
#
# Column boundaries below were re-measured directly against a real HDFC
# "BIZ PRO PLUS ACCOUNT" statement (612pt-wide page — NOT 638pt as the
# old comment assumed). The previous boundaries were shifted right by
# roughly 15-25pt across every column, which caused two silent failures:
#
#   1. Chq./Ref.No. column real x0 ≈ 268-337, but the old 'ref_no' band
#      was (290, 367) — that overlapped the *Value Dt* column (real x0
#      ≈ 346.7), so `ref_no` was actually getting filled with the value
#      date string, while the real reference number fell inside the old
#      'narration' band (68, 290) and got silently appended to the
#      description instead.
#   2. Withdrawal/Deposit/Balance bands were each shifted right by a
#      similar margin, which is harmless only because amounts happen to
#      be right-aligned within their real columns — but any statement
#      with wider (more digits) amounts than this test file could have
#      pushed a value's x0 left into the gap between the old 'value_date'
#      and 'withdrawal' bands (362-405), where it would silently vanish
#      (parsed as neither a date nor an amount).
#
# New boundaries (verified with pdfplumber word coordinates across all
# transaction rows of a real statement, x_tolerance=2):
#   Date         x0 ≈ 38.5           → band (0, 68)
#   Narration    x0 ≈ 70-218         → band (68, 268)
#   Chq./Ref.No. x0 ≈ 268-337        → band (268, 344)
#   Value Dt     x0 ≈ 346.7 (fixed)  → band (344, 386)
#   Withdrawal   x0 ≈ 386-448        → band (386, 462)
#   Deposit      x0 ≈ 487-521        → band (462, 532)
#   Closing Bal  x0 ≈ 560-594        → band (532, 612)
# ─────────────────────────────────────────────────────────────────────────────



_HDFC_HEADER_ALIASES = {
    "date": [
        ["date"],
        ["txn", "date"],
        ["transaction", "date"],
    ],

    "narration": [
        ["narration"],
        ["description"],
        ["particulars"],
    ],

    "ref_no": [
        ["chq", "ref", "no"],
        ["chq", "ref"],
        ["ref", "no"],
        ["reference", "no"],
        ["reference"],
    ],

    "value_date": [
        ["value", "dt"],
        ["value", "date"],
    ],

    "withdrawal": [
        ["withdrawal"],
        ["withdrawal", "amt"],
        ["debit"],
        ["debit", "amt"],
    ],

    "deposit": [
        ["deposit"],
        ["deposit", "amt"],
        ["credit"],
        ["credit", "amt"],
    ],

    "balance": [
        ["closing", "balance"],
        ["balance"],
        ["closing", "bal"],
    ],
}

# _HDFC_COLS = {
#     'date':       (0,   68),
#     'narration':  (68,  268),
#     'ref_no':     (268, 344),
#     'value_date': (344, 386),
#     'withdrawal': (386, 462),
#     'deposit':    (462, 532),
#     'balance':    (532, 612),
# }
_HDFC_DATE_FMTS = ['%d/%m/%y', '%d/%m/%Y']


def _normalize_hdfc_header(text: str) -> str:
    """
    Generic normalization only.

    Does not contain bank-format-specific mappings.
    """

    text = str(text or "").strip()

    # Normalize CamelCase / concatenated PDF words.
    # Examples:
    #   ValueDt         -> Value Dt
    #   WithdrawalAmt   -> Withdrawal Amt
    #   ClosingBalance  -> Closing Balance
    text = re.sub(
        r"(?<=[a-z])(?=[A-Z])",
        " ",
        text,
    )

    # Normalize punctuation/separators.
    text = re.sub(
        r"[^a-zA-Z0-9]+",
        " ",
        text,
    )

    # Normalize repeated whitespace.
    text = re.sub(
        r"\s+",
        " ",
        text,
    ).strip()

    return text.lower()


def _tokenize_hdfc_header(text: str) -> list[str]:
    """
    Convert a header string into semantic tokens.

    Examples:

        ValueDt
            -> ["value", "dt"]

        WithdrawalAmt.
            -> ["withdrawal", "amt"]

        ClosingBalance
            -> ["closing", "balance"]

        Chq./Ref.No.
            -> ["chq", "ref", "no"]
    """

    normalized = _normalize_hdfc_header(text)

    if not normalized:
        return []

    return normalized.split()


def _match_hdfc_header_alias(
    row: list[dict],
    start_index: int,
    alias: list[str],
) -> Optional[dict]:
    """
    Match a semantic HDFC header alias against PDF words.

    Matching is token-based rather than exact-string based.

    Example:

        PDF:
            WithdrawalAmt.

        Tokens:
            ["withdrawal", "amt"]

        Alias:
            ["withdrawal", "amt"]

        Result:
            MATCH
    """

    if not alias:
        return None

    collected_words = []
    collected_tokens = []

    index = start_index

    while index < len(row):

        word = row[index]

        tokens = _tokenize_hdfc_header(
            word.get("text", "")
        )

        if not tokens:
            index += 1
            continue

        collected_words.append(word)
        collected_tokens.extend(tokens)

        if collected_tokens == alias:

            return {
                "x0": min(
                    w["x0"]
                    for w in collected_words
                ),
                "x1": max(
                    w["x1"]
                    for w in collected_words
                ),
                "start_index": start_index,
                "end_index": index,
                "words": collected_words,
            }

        if len(collected_tokens) >= len(alias):
            break

        index += 1

    return None


def _debug_hdfc_header_rows(rows: dict, page_num: int) -> None:
    """
    Temporary diagnostic helper.

    Prints the actual words and coordinates that pdfplumber
    extracted from the HDFC page so we can see why header
    detection is failing.
    """

    if page_num != 1:
        return

    logger.warning(
        "========== HDFC HEADER DEBUG PAGE %d ==========",
        page_num,
    )

    for row_number, top in enumerate(sorted(rows), 1):

        row = sorted(
            rows[top],
            key=lambda w: w['x0'],
        )

        if not row:
            continue

        row_text = " | ".join(
            str(w['text'])
            for w in row
        )

        logger.warning(
            "HDFC ROW %d | top=%s | %s",
            row_number,
            top,
            row_text,
        )

        for word in row:
            logger.warning(
                "    text=%r x0=%.2f x1=%.2f top=%.2f bottom=%.2f",
                word.get('text'),
                word.get('x0', 0),
                word.get('x1', 0),
                word.get('top', 0),
                word.get('bottom', 0),
            )

    logger.warning(
        "========== END HDFC HEADER DEBUG =========="
    )




def _detect_hdfc_header(rows: dict) -> Optional[dict]:
    """
    Detect the logical HDFC transaction header.

    Supports:
        1. A single header row.
        2. A two-row visual header.

    The resulting header contains the actual PDF coordinates
    of each semantic column.
    """

    row_items = []

    for top in sorted(rows):

        row = sorted(
            rows[top],
            key=lambda w: w['x0']
        )

        if not row:
            continue

        row_items.append({
            'top': top,
            'words': row,
        })

    candidates = []

    # ---------------------------------------------------------
    # Try a single-row header and a two-row header.
    # ---------------------------------------------------------

    for index in range(len(row_items)):

        # -----------------------------------------------------
        # Candidate 1: single visual row
        # -----------------------------------------------------

        windows = [
            [row_items[index]]
        ]

        # -----------------------------------------------------
        # Candidate 2: two nearby visual rows
        #
        # We only merge immediately adjacent rows. We do not
        # scan arbitrary rows across the page.
        # -----------------------------------------------------

        if index + 1 < len(row_items):

            current = row_items[index]
            next_row = row_items[index + 1]

            vertical_gap = (
                next_row['top']
                - current['top']
            )

            # Header lines should be close together.
            #
            # 15 points is intentionally conservative so that
            # we don't accidentally merge unrelated statement
            # sections.
            if 0 < vertical_gap <= 15:

                windows.append([
                    current,
                    next_row,
                ])

        for window in windows:

            # -------------------------------------------------
            # Flatten the words from the visual header rows.
            #
            # Keep physical coordinates from the PDF.
            # -------------------------------------------------

            logical_words = []

            for row_info in window:

                for word in row_info['words']:

                    logical_words.append(word)

            if not logical_words:
                continue

            # -------------------------------------------------
            # For semantic matching, order by physical position.
            #
            # For words on different header lines, Y is also
            # considered so that their original reading order
            # remains stable.
            # -------------------------------------------------

            logical_words.sort(
                key=lambda w: (
                    round(w['top'], 2),
                    w['x0'],
                )
            )

            detected = {}

            # -------------------------------------------------
            # Search the complete logical header.
            # -------------------------------------------------

            for start_index in range(
                len(logical_words)
            ):

                for field, aliases in (
                    _HDFC_HEADER_ALIASES.items()
                ):

                    if field in detected:
                        continue

                    ordered_aliases = sorted(
                        aliases,
                        key=len,
                        reverse=True,
                    )

                    for alias in ordered_aliases:

                        match = _match_hdfc_header_alias(
                            row=logical_words,
                            start_index=start_index,
                            alias=alias,
                        )

                        if match:

                            detected[field] = {
                                'text': alias,
                                'x0': match['x0'],
                                'x1': match['x1'],
                                'start_index': (
                                    match['start_index']
                                ),
                                'end_index': (
                                    match['end_index']
                                ),
                                'words': match['words'],
                            }

                            break

                    if field in detected:
                        continue

            # -------------------------------------------------
            # Validate that this is actually a transaction
            # header.
            # -------------------------------------------------

            core_fields = {
                'date',
                'narration',
                'balance',
            }

            amount_fields = {
                'withdrawal',
                'deposit',
            }

            core_match = (
                core_fields.intersection(
                    detected.keys()
                )
            )

            amount_match = (
                amount_fields.intersection(
                    detected.keys()
                )
            )

            score = (
                len(core_match) * 3
                + len(amount_match) * 2
                + (
                    1
                    if 'ref_no' in detected
                    else 0
                )
                + (
                    1
                    if 'value_date' in detected
                    else 0
                )
            )

            if (
                len(core_match) == len(core_fields)
                and amount_match
            ):
                candidates.append({
                    'row_top': window[0]['top'],
                    'row_bottom': window[-1]['top'],
                    'words': logical_words,
                    'columns': detected,
                    'score': score,
                    'row_count': len(window),
                })

    if not candidates:
        return None

    # ---------------------------------------------------------
    # Prefer:
    #
    # 1. Highest semantic score.
    # 2. Single-row header when scores are equal.
    #
    # This prevents us from unnecessarily merging two rows.
    # ---------------------------------------------------------

    candidates.sort(
        key=lambda item: (
            item['score'],
            -item['row_count'],
        ),
        reverse=True,
    )

    return candidates[0]


def _build_hdfc_columns(
    header: dict,
    rows: dict,
) -> Optional[dict]:
    """
    Build dynamic HDFC transaction-column boundaries.

    Header coordinates are used as semantic anchors.
    They are NOT assumed to be the actual left boundary
    of the transaction column.

    Transaction-row X coordinates are used to determine
    the actual content area dynamically.
    """

    detected = header.get('columns', {})

    if not detected:
        return None

    # ---------------------------------------------------------
    # Sort detected columns according to their header position.
    # ---------------------------------------------------------

    ordered = sorted(
        detected.items(),
        key=lambda item: item[1]['x0'],
    )

    if not ordered:
        return None

    # ---------------------------------------------------------
    # Detect transaction-like rows.
    #
    # We use the detected DATE header as the anchor for finding
    # rows that actually contain transactions.
    # ---------------------------------------------------------

    transaction_rows = []

    date_info = detected.get('date')

    if date_info:

        date_x0 = date_info['x0']
        date_x1 = date_info['x1']

        for row in rows.values():

            if not row:
                continue

            has_date = False

            for word in row:

                word_x0 = word.get('x0')

                if word_x0 is None:
                    continue

                # Allow some tolerance because the transaction
                # date does not necessarily have the same x0
                # as the header text.
                if not (
                    date_x0 - 30
                    <= word_x0
                    <= date_x1 + 30
                ):
                    continue

                if _parse_date(
                    word.get('text', ''),
                    _HDFC_DATE_FMTS,
                ):
                    has_date = True
                    break

            if has_date:
                transaction_rows.append(row)

    # ---------------------------------------------------------
    # If no transaction rows were found, we cannot safely infer
    # dynamic transaction boundaries.
    #
    # Do not create fixed/fallback coordinates.
    # ---------------------------------------------------------

    if not transaction_rows:
        return None

    # ---------------------------------------------------------
    # For every detected semantic column, find the transaction
    # words that are horizontally closest to the corresponding
    # header anchor.
    #
    # This gives us the real transaction-side X position.
    # ---------------------------------------------------------

    transaction_anchors = {}

    for field, info in ordered:

        header_x0 = info['x0']
        header_x1 = info['x1']

        header_center = (
            header_x0 + header_x1
        ) / 2

        candidates = []

        for row in transaction_rows:

            for word in row:

                word_x0 = word.get('x0')
                word_x1 = word.get('x1')

                if (
                    word_x0 is None
                    or word_x1 is None
                ):
                    continue

                word_center = (
                    word_x0 + word_x1
                ) / 2

                distance = abs(
                    word_center - header_center
                )

                candidates.append(
                    (
                        distance,
                        word_x0,
                        word_x1,
                    )
                )

        if not candidates:
            continue

        # -----------------------------------------------------
        # Use the transaction word closest to the header anchor.
        # -----------------------------------------------------

        candidates.sort(
            key=lambda item: item[0]
        )

        _, anchor_x0, anchor_x1 = candidates[0]

        transaction_anchors[field] = {
            'x0': anchor_x0,
            'x1': anchor_x1,
        }

    # ---------------------------------------------------------
    # We need all detected columns to build a reliable ordered
    # set of transaction anchors.
    # ---------------------------------------------------------

    if len(transaction_anchors) != len(ordered):
        return None

    # ---------------------------------------------------------
    # Sort columns according to their transaction anchors.
    # ---------------------------------------------------------

    ordered_transaction = sorted(
        transaction_anchors.items(),
        key=lambda item: item[1]['x0'],
    )

    columns = {}

    # ---------------------------------------------------------
    # Build boundaries between neighboring transaction anchors.
    #
    # The boundary is halfway between the actual transaction
    # positions, NOT halfway between header text positions.
    # ---------------------------------------------------------

    for index, (field, anchor) in enumerate(
        ordered_transaction
    ):

        anchor_x0 = anchor['x0']
        anchor_x1 = anchor['x1']

        # -----------------------------------------------------
        # First column.
        #
        # We deliberately extend to the left of the detected
        # transaction word so that different transaction text
        # widths do not get cut off.
        # -----------------------------------------------------

        if index == 0:

            column_start = 0.0

        else:

            previous_field, previous_anchor = (
                ordered_transaction[index - 1]
            )

            previous_x1 = previous_anchor['x1']

            column_start = (
                previous_x1 + anchor_x0
            ) / 2

        # -----------------------------------------------------
        # Last column.
        # -----------------------------------------------------

        if index == len(
            ordered_transaction
        ) - 1:

            max_x1 = max(
                word['x1']
                for row in transaction_rows
                for word in row
                if word.get('x1') is not None
            )

            column_end = max(
                max_x1 + 20,
                anchor_x1 + 20,
            )

        else:

            _, next_anchor = (
                ordered_transaction[index + 1]
            )

            next_x0 = next_anchor['x0']

            column_end = (
                anchor_x1 + next_x0
            ) / 2

        # -----------------------------------------------------
        # Ensure the boundary is valid.
        # -----------------------------------------------------

        if column_start >= column_end:
            return None

        columns[field] = (
            column_start,
            column_end,
        )

    return columns


def _validate_hdfc_columns(
    header: dict,
    columns: dict,
) -> bool:

    required_fields = {
        'date',
        'narration',
        'balance',
    }

    amount_fields = {
        'withdrawal',
        'deposit',
    }

    if not required_fields.issubset(columns):
        return False

    if not amount_fields.intersection(columns):
        return False

    for field, bounds in columns.items():

        if not isinstance(bounds, tuple):
            return False

        if len(bounds) != 2:
            return False

        x0, x1 = bounds

        if x0 >= x1:
            return False

    ordered = sorted(
        columns.items(),
        key=lambda item: item[1][0],
    )

    for index in range(1, len(ordered)):

        previous_x1 = ordered[index - 1][1][1]
        current_x0 = ordered[index][1][0]

        if current_x0 < previous_x1:
            return False

    positions = {
        field: bounds[0]
        for field, bounds in columns.items()
    }

    if positions['date'] >= positions['narration']:
        return False

    if 'ref_no' in positions:
        if positions['ref_no'] <= positions['narration']:
            return False

    if 'value_date' in positions:

        previous_field = (
            'ref_no'
            if 'ref_no' in positions
            else 'narration'
        )

        if positions['value_date'] <= positions[previous_field]:
            return False

    amount_positions = [
        positions[field]
        for field in amount_fields
        if field in positions
    ]

    if not amount_positions:
        return False

    if positions['balance'] <= max(amount_positions):
        return False

    # detected = header.get('columns', {})

    # for field in columns:

    #     if field not in detected:
    #         return False

    #     if abs(
    #         detected[field]['x0']
    #         - columns[field][0]
    #     ) > 0.01:
    #         return False

    return True


def _in_hdfc(
    word: dict,
    col: str,
    columns: dict,
) -> bool:

    if col not in columns:
        return False

    x0, x1 = columns[col]

    return x0 <= word['x0'] < x1




def _parse_hdfc(pdf) -> tuple:
    txns, errors = [], []

    detected_columns = None

    for page_num, page in enumerate(pdf.pages, 1):
        try:
            words = page.extract_words(
                x_tolerance=2,
                y_tolerance=3,
            )

            rows: dict = defaultdict(list)

            for w in words:
                rows[round(w['top'])].append(w)

            # -------------------------------------------------
            # TEMPORARY HDFC HEADER DEBUG
            # -------------------------------------------------

            _debug_hdfc_header_rows(
                rows,
                page_num,)
            # -------------------------------------------------
            # Detect HDFC header from this page.
            # -------------------------------------------------

            header = _detect_hdfc_header(rows)

            if header:
                logger.warning(
                    "HDFC HEADER DETECTED | score=%s | row_count=%s | columns=%s",
                    header.get("score"),
                    header.get("row_count"),
                    list(header.get("columns", {}).keys()),
                )
            else:
                logger.warning(
                    "HDFC HEADER NOT DETECTED"
                )

            if header:
                page_columns = _build_hdfc_columns(header, rows)

                if not page_columns:
                    raise ValueError(
                        'HDFC header detected but column '
                        'coordinates could not be built'
                    )

                if not _validate_hdfc_columns(
                    header,
                    page_columns,
                ):
                    raise ValueError(
                        'HDFC detected column structure '
                        'failed validation'
                    )

                detected_columns = page_columns

            # -------------------------------------------------
            # No detected coordinates means we cannot safely
            # parse this statement.
            #
            # There is intentionally NO fixed-coordinate
            # fallback here.
            # -------------------------------------------------

            if detected_columns is None:
                raise ValueError(
                    'HDFC transaction header could not '
                    'be reliably detected'
                )

            pending = None

            for top in sorted(rows):

                rw = sorted(
                    rows[top],
                    key=lambda w: w['x0']
                )

                if not rw:
                    continue

                first = rw[0]

                is_txn = (
                    _in_hdfc(
                        first,
                        'date',
                        detected_columns,
                    )
                    and bool(
                        re.match(
                            r'^\d{2}/\d{2}/\d{2,4}$',
                            first['text'],
                        )
                    )
                )

                if is_txn:

                    if pending:
                        t = _build_hdfc_txn(pending)

                        if t:
                            txns.append(t)

                    pending = {
                        'date_str': first['text'],
                        'narration': [],
                        'ref': '',
                        'value_date': None,
                        'withdrawal': None,
                        'deposit': None,
                        'balance': None,
                    }

                    for w in rw[1:]:

                        txt = w['text']

                        if _in_hdfc(
                            w,
                            'narration',
                            detected_columns,
                        ):
                            pending['narration'].append(txt)

                        elif _in_hdfc(
                            w,
                            'ref_no',
                            detected_columns,
                        ):
                            pending['ref'] = txt

                        elif _in_hdfc(
                            w,
                            'value_date',
                            detected_columns,
                        ):
                            pending['value_date'] = _parse_date(
                                txt,
                                _HDFC_DATE_FMTS,
                            )

                        elif _in_hdfc(
                            w,
                            'withdrawal',
                            detected_columns,
                        ):
                            a = _parse_amount_positive(txt)

                            if a:
                                pending['withdrawal'] = a

                        elif _in_hdfc(
                            w,
                            'deposit',
                            detected_columns,
                        ):
                            a = _parse_amount_positive(txt)

                            if a:
                                pending['deposit'] = a

                        elif _in_hdfc(
                            w,
                            'balance',
                            detected_columns,
                        ):
                            a = _parse_amount_positive(txt)

                            if a:
                                pending['balance'] = a

                else:

                    # Continuation row.
                    if pending:

                        for w in rw:

                            txt = w['text']

                            if _in_hdfc(
                                w,
                                'narration',
                                detected_columns,
                            ):
                                pending['narration'].append(txt)

                            elif (
                                _in_hdfc(
                                    w,
                                    'withdrawal',
                                    detected_columns,
                                )
                                and pending['withdrawal'] is None
                            ):
                                a = _parse_amount_positive(txt)

                                if a:
                                    pending['withdrawal'] = a

                            elif (
                                _in_hdfc(
                                    w,
                                    'deposit',
                                    detected_columns,
                                )
                                and pending['deposit'] is None
                            ):
                                a = _parse_amount_positive(txt)

                                if a:
                                    pending['deposit'] = a

                            elif (
                                _in_hdfc(
                                    w,
                                    'balance',
                                    detected_columns,
                                )
                                and pending['balance'] is None
                            ):
                                a = _parse_amount_positive(txt)

                                if a:
                                    pending['balance'] = a

            if pending:

                t = _build_hdfc_txn(pending)

                if t:
                    txns.append(t)

        except Exception as e:

            errors.append(
                f'HDFC page {page_num}: {e}'
            )

            logger.warning(
                'HDFC page %d error: %s',
                page_num,
                e,
            )

    return txns, errors

def _build_hdfc_txn(p: dict) -> Optional[ParsedTransaction]:
    txn_date = _parse_date(p['date_str'], _HDFC_DATE_FMTS)
    if not txn_date:
        return None

    wd  = p.get('withdrawal')
    dep = p.get('deposit')

    if wd and wd > 0:
        txn_type, amount = 'debit', wd
    elif dep and dep > 0:
        txn_type, amount = 'credit', dep
    else:
        return None

    desc = _clean_description(' '.join(p['narration']))
    ref  = p.get('ref', '')
    utr  = _extract_utr(desc + ' ' + ref)

    return ParsedTransaction(
        txn_date=txn_date,
        value_date=p.get('value_date'),
        description=desc,
        ref_no=ref,
        utr_no=utr,
        txn_type=txn_type,
        amount=amount,
        balance=p.get('balance'),
    )


# ─────────────────────────────────────────────────────────────────────────────
# Generic table parser  (ICICI, AXIS, unknown banks)
# ─────────────────────────────────────────────────────────────────────────────

_GENERIC_DATE_FMTS = [
    '%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y', '%d-%m-%y',
    '%d-%b-%Y', '%d-%b-%y', '%d %b %Y',
]

_GENERIC_HDR = {
    'date': 'date', 'txn date': 'date', 'transaction date': 'date',
    'value date': 'vdate',
    'description': 'desc', 'narration': 'desc', 'particulars': 'desc',
    'ref no': 'ref', 'cheque': 'ref', 'reference': 'ref',
    'debit': 'debit', 'withdrawal': 'debit', 'dr': 'debit',
    'credit': 'credit', 'deposit': 'credit', 'cr': 'credit',
    'balance': 'balance', 'closing balance': 'balance',
}


def _parse_generic(pdf) -> tuple:
    txns, errors = [], []

    for page_num, page in enumerate(pdf.pages, 1):
        try:
            tables = page.extract_tables()
            for table in tables:
                if not table or len(table) < 2:
                    continue

                hdr_idx = None
                for i, row in enumerate(table):
                    joined = ' '.join(str(c) for c in row if c).lower()
                    score = sum(1 for k in ['date', 'debit', 'credit', 'balance'] if k in joined)
                    if score >= 2:
                        hdr_idx = i
                        break

                start = (hdr_idx + 1) if hdr_idx is not None else 0
                headers = [str(c or '').strip().lower() for c in table[hdr_idx or 0]]
                col: dict = {}
                for i, h in enumerate(headers):
                    for kw, fld in _GENERIC_HDR.items():
                        if kw in h and fld not in col:
                            col[fld] = i
                            break

                for row in table[start:]:
                    if not row:
                        continue
                    date_raw = str(row[col.get('date', 0)] or '').split('\n')[0].strip()
                    txn_date = _parse_date(date_raw, _GENERIC_DATE_FMTS)
                    if not txn_date:
                        continue

                    desc  = _clean_description(str(row[col.get('desc', 1)] or '').replace('\n', ' '))
                    ref   = str(row[col.get('ref', 2)] or '').replace('\n', ' ').strip()
                    debit  = _parse_amount_positive(row[col['debit']])  if 'debit'  in col else None
                    credit = _parse_amount_positive(row[col['credit']]) if 'credit' in col else None
                    bal    = _parse_amount(row[col['balance']])          if 'balance' in col else None

                    if debit:
                        txn_type, amount = 'debit', debit
                    elif credit:
                        txn_type, amount = 'credit', credit
                    else:
                        continue

                    utr = _extract_utr(desc + ' ' + ref)
                    txns.append(ParsedTransaction(
                        txn_date=txn_date,
                        value_date=None,
                        description=desc,
                        ref_no=ref,
                        utr_no=utr,
                        txn_type=txn_type,
                        amount=amount,
                        balance=bal,
                    ))

        except Exception as e:
            errors.append(f'Generic page {page_num}: {e}')

    return txns, errors


# ─────────────────────────────────────────────────────────────────────────────
# Main parser class
# ─────────────────────────────────────────────────────────────────────────────

class BankStatementParser:

    def parse_pdf(self, file_bytes: bytes, file_name: str = "") -> ParseResult:
        try:
            if _is_excel_file(file_name):
                full_text = _extract_excel_text(file_bytes)
                ai_result = _ollama_statement_result(full_text)
                if ai_result:
                    logger.info(
                        'Parsed %d transactions from Excel statement using Ollama',
                        len(ai_result.transactions),
                    )
                    return ai_result
                return ParseResult(
                    bank_name='UNKNOWN',
                    account_no='',
                    account_name='',
                    errors=['No transactions extracted from Excel statement'],
                )

            if _is_image_file(file_name):
                full_text = _extract_image_text(file_bytes)
                ai_result = _ollama_statement_result(full_text)
                if ai_result:
                    logger.info(
                        'Parsed %d transactions from image statement using Ollama',
                        len(ai_result.transactions),
                    )
                    return ai_result
                return ParseResult(
                    bank_name='UNKNOWN',
                    account_no='',
                    account_name='',
                    errors=['No transactions extracted from image statement'],
                )

            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                full_text = '\n'.join(
                    page.extract_text(x_tolerance=2, y_tolerance=2) or ''
                    for page in pdf.pages
                )
                if len(full_text.strip()) < 100:
                    full_text = _extract_pdf_text_with_ocr(file_bytes) or full_text

                bank = _detect_bank(full_text)
                acc_no, acc_name = _extract_metadata(full_text)

                if bank == 'DBS':
                    txns, errors = _parse_dbs(pdf)
                elif bank == 'SBI':
                    txns, errors = _parse_sbi(pdf)
                elif bank == 'HDFC':
                    txns, errors = _parse_hdfc(pdf)
                else:
                    # ICICI, AXIS, unknown — generic table parser
                    txns, errors = _parse_generic(pdf)
                    if not txns:
                        # second chance: try all specific parsers
                        for fn in [_parse_sbi, _parse_hdfc, _parse_dbs]:
                            txns, errs = fn(pdf)
                            if txns:
                                break
                        else:
                            errors.append('No transactions extracted by any parser')

                txns = _deduplicate(txns)

                ai_result = _ollama_statement_result(full_text)
                if ai_result and (
                    not txns or
                    (len(ai_result.transactions) > len(txns) and bank == 'UNKNOWN')
                ):
                    logger.info(
                        'Using Ollama statement fallback: %d transactions from %s',
                        len(ai_result.transactions),
                        ai_result.bank_name,
                    )
                    return ai_result

                logger.info('Parsed %d transactions from %s statement', len(txns), bank)

                return ParseResult(
                    bank_name=bank,
                    account_no=acc_no,
                    account_name=acc_name,
                    transactions=txns,
                    errors=errors,
                )

        except Exception as e:
            image_text = _extract_image_text(file_bytes)
            ai_result = _ollama_statement_result(image_text)
            if ai_result:
                logger.info(
                    'Parsed %d transactions from statement image fallback using Ollama',
                    len(ai_result.transactions),
                )
                return ai_result
            logger.error('PDF parsing failed: %s', e, exc_info=True)
            return ParseResult(
                bank_name='ERROR',
                account_no='',
                account_name='',
                errors=[f'PDF parsing failed: {e}'],
            )


# ─────────────────────────────────────────────────────────────────────────────
# Public API (backward-compatible)
# ─────────────────────────────────────────────────────────────────────────────

def parse_bank_statement(file_bytes: bytes, file_name: str = "") -> ParseResult:
    """Entry point called by views.py — unchanged signature."""
    return BankStatementParser().parse_pdf(file_bytes, file_name)