# banklist/views.py

import io
import logging
from decimal import Decimal
from datetime import datetime, date
from django.db.models import Q, Count, Sum, Prefetch
from django.utils import timezone
from django.core.cache import cache

from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework_simplejwt.tokens import RefreshToken

from celery.result import AsyncResult
import re

from .models import (
    Bank, BankAccount, Transaction, UploadedStatement,
    ReceiptDocument, Company, User, ReconciliationRun,
    ReconciliationRule, AuditLog, Particular, AccountingParticular, Grant, Fund, TransactionFundAllocation, GrantMaster, InternalTransferTransaction,
    InternalTransfer, Agency, GrantTransaction, GrantMilestone, TransactionGrantAllocation
)
from .serializers import (
    BankSerializer, BankAccountSerializer, TransactionSerializer,
    UploadedStatementSerializer, ReceiptDocumentSerializer,
    ReconciliationRunSerializer, ReconciliationRuleSerializer,
    AuditLogSerializer, RegisterSerializer, LoginSerializer, ParticularSerializer, 
    TransactionUpdateSerializer, AccountingParticularDropdownSerializer, TransactionJourneySerializer,
    FundSerializer, GrantListSerializer, ListAgencySerializer, GrantSerializer,
    GrantMilestoneSerializer, GrantTransactionSerializer, CreateAgencySerializer
)
from .services.google_drive_service import drive_service
from .services.pdf_parser import parse_bank_statement
from .services.receipt_parser import parse_receipt_pdf, extract_receipt_data
from .services.reconciliation_engine import run_reconciliation
from .services.enhanced_reconciliation_engine import (
    run_full_reconciliation,
    match_with_receipts_only
)
from .tasks import (
    create_drive_folder_task,
    upload_to_drive_task,
    parse_statement_task,
    process_receipt_task,
    process_all_receipts_task,
    run_reconciliation_task
)
from banklist.services.reconciliation_v2_engine import (
    run_reconciliation_v2,
    auto_receipt_matching
)
from django.shortcuts import get_object_or_404
import calendar

from .services.report_service import (
    # get_week_ranges,
    # generate_cash_flow_report,
    CashFlowAnalysisService,
    GrantWiseOutflowService
)
from django.db import transaction
from banklist.services.fund_service import get_grant_available_amount
from decimal import Decimal, InvalidOperation
from django.db import transaction as db_transaction
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse


logger = logging.getLogger(__name__)


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def _get_company_name(request):
    try:
        user = request.user
        if user.company:
            return user.company.name
        return user.full_name or user.email
    except Exception as e:
        logger.error(f"Error getting company name: {e}")
        return None


def _parse_drive_timestamp(ts: str):
    """
    Parse Google Drive ISO timestamp: '2026-06-28T10:20:59.000Z'
    Returns a timezone-aware datetime or None.

    FIX: views.py was calling _parse_date() on this format which
    only handles dd/mm/yyyy style — always returned None.
    """
    if not ts:
        return None
    try:
        # Strip milliseconds if present
        ts_clean = re.sub(r'\.\d+Z$', 'Z', ts)
        dt = datetime.strptime(ts_clean, '%Y-%m-%dT%H:%M:%SZ')
        return timezone.make_aware(dt, timezone.utc)
    except Exception:
        return None


def _parse_date(date_str):
    """Parse dd/mm/yyyy style date strings (used for bank statement dates)."""
    if not date_str:
        return None
    date_str = str(date_str).strip()
    formats = [
        '%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y', '%d-%m-%y',
        '%d-%b-%Y', '%d-%b-%y', '%d %b %Y', '%d %b %y',
        '%Y-%m-%d', '%Y/%m/%d', '%d.%m.%Y', '%d.%m.%y'
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def _extract_receipt_text(file_bytes, mime_type):
    """Extract text from receipt with OCR fallback."""
    if not file_bytes:
        return ""
    try:
        import pdfplumber
        import fitz
        from PIL import Image
        import pytesseract

        extracted_text = ""
        if mime_type == "application/pdf":
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                extracted_text = "\n".join(
                    page.extract_text() or "" for page in pdf.pages
                ).strip()

        ocr_needed = (
            mime_type != "application/pdf" or
            len(extracted_text) < 200 or
            not re.search(r"(₹|INR|Rs\.?)", extracted_text, re.I)
        )

        if not ocr_needed:
            return extracted_text

        if mime_type == "application/pdf":
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            pages = []
            for page in doc:
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
                img = Image.open(io.BytesIO(pix.tobytes("png")))
                pages.append(pytesseract.image_to_string(img, lang="eng"))
            return "\n".join(pages).strip()

        return pytesseract.image_to_string(
            Image.open(io.BytesIO(file_bytes)), lang="eng"
        ).strip()

    except Exception as e:
        logger.error(f"Receipt text extraction failed: {e}")
        return ""


# ============================================================
# AUTHENTICATION VIEWS
# ============================================================

@api_view(['POST'])
@permission_classes([AllowAny])
def register(request):
    print(request.data)
    serializer = RegisterSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    user = serializer.save()
    refresh = RefreshToken.for_user(user)

    return Response({
        'success': True,
        'message': 'Account created successfully',
        'user': {
            'id': user.id,
            'email': user.email,
            'full_name': user.full_name,
            'company': user.company.name,
        },
        'tokens': {
            'access': str(refresh.access_token),
            'refresh': str(refresh),
        }
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    print(request.data)
    serializer = LoginSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    user = serializer.validated_data['user']
    refresh = RefreshToken.for_user(user)

    return Response({
        'success': True,
        'message': 'Login successful',
        'user': {
            'id': user.id,
            'email': user.email,
            'full_name': user.full_name,
            'company': user.company.name if user.company else None,
        },
        'tokens': {
            'access': str(refresh.access_token),
            'refresh': str(refresh),
        }
    }, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([AllowAny])
def logout(request):
    try:
        token = RefreshToken(request.data.get('refresh'))
        token.blacklist()
        return Response({'success': True, 'message': 'Logged out'})
    except Exception:
        return Response({'error': 'Invalid token'}, status=status.HTTP_400_BAD_REQUEST)


# ============================================================
# BANK MANAGEMENT VIEWS
# ============================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def search_banks(request):
    search_term = request.GET.get('search', '').strip()
    banks = (
        Bank.objects.filter(bank_name__icontains=search_term, is_active=True)
        if search_term else
        Bank.objects.filter(is_active=True)
    )
    return Response([
        {'bank_name': b.bank_name, 'short_name': b.short_name, 'id': b.id}
        for b in banks
    ])


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_bank_accounts(request):
    company = request.user.company
    if not company:
        return Response({'error': 'User has no company'}, status=400)
    accounts = BankAccount.objects.filter(company=company).select_related('bank')
    return Response(BankAccountSerializer(accounts, many=True).data)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_bank_account(request, account_id):
    try:
        bank_account = BankAccount.objects.get(
            id=account_id, company=request.user.company
        )
    except BankAccount.DoesNotExist:
        return Response({'error': 'Bank account not found'}, status=404)

    try:
        bank_name    = str(bank_account)
        drive_errors = []

        if bank_account.bank_folder_id:
            try:
                if not drive_service.delete_folder(bank_account.bank_folder_id):
                    drive_errors.append('Could not delete Drive folder')
            except Exception as e:
                logger.error(f"Drive delete error: {e}")
                drive_errors.append(str(e))

        bank_account.delete()

        return Response({
            'success': True,
            'message': f'"{bank_name}" deleted successfully',
            'drive_warnings': drive_errors,
        })

    except Exception as e:
        logger.error(f"Delete error: {e}", exc_info=True)
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def init_company_folder(request):
    company_name = _get_company_name(request)
    if not company_name:
        return Response({'error': 'Could not determine company/user name'}, status=400)

    result = drive_service.get_or_create_company_folder(company_name)
    if not result:
        return Response({'error': 'Failed to create company folder'}, status=500)

    return Response({
        'success': True,
        'message': f'Company folder ready: {company_name}',
        'company_name': company_name,
        'company_folder_id': result['company_folder_id'],
        'subfolders': result['subfolders'],
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_drive_folder(request):
    bank_name      = request.data.get('bank_name')
    account_holder = request.data.get('account_holder_name')
    account_number = request.data.get('account_number')
    ifsc_code      = request.data.get('ifsc_code', '')

    if not bank_name:
        return Response({'error': 'bank_name is required'}, status=400)

    bank = Bank.objects.filter(bank_name=bank_name).first()
    if not bank:
        return Response({'error': f'Bank {bank_name} not found'}, status=404)

    folder_name  = f"{bank_name} - {account_holder} - {account_number}"
    company_name = _get_company_name(request)

    drive_result = drive_service.create_bank_folder_structure(
        folder_name, company_name=company_name
    )
    if not drive_result['success']:
        return Response({'error': drive_result['message']}, status=500)

    bank_account, _ = BankAccount.objects.get_or_create(
        company=request.user.company,
        bank=bank,
        account_number=account_number,
        defaults={
            'account_holder_name': account_holder,
            'ifsc_code':           ifsc_code,
            'bank_folder_id':      drive_result.get('bank_folder_id'),
            'statement_folder_id': drive_result.get('statement_folder_id'),
            'drive_link':          drive_result.get('bank_folder_link'),
        }
    )

    return Response({
        'success': True,
        'message': f'Google Drive folders created for {folder_name}',
        'bank_account': BankAccountSerializer(bank_account).data,
        'drive_folders': {'total_folders': drive_result.get('total_folders_created', 0)},
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_drive_folder_async(request):
    bank_name      = request.data.get('bank_name')
    account_holder = request.data.get('account_holder_name')
    account_number = request.data.get('account_number')

    if not bank_name:
        return Response({'error': 'bank_name is required'}, status=400)
    if not Bank.objects.filter(bank_name=bank_name).exists():
        return Response({'error': f'Bank {bank_name} not found'}, status=404)

    folder_name  = f"{bank_name} - {account_holder} - {account_number}"
    company_name = _get_company_name(request)
    task         = create_drive_folder_task.delay(folder_name, company_name=company_name)

    return Response({
        'success':    True,
        'message':    'Folder creation queued',
        'task_id':    task.id,
        'status_url': f"/api/tasks/{task.id}/status/",
    }, status=status.HTTP_202_ACCEPTED)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def task_status(request, task_id):
    task_result = AsyncResult(task_id)
    response = {
        'task_id':    task_id,
        'status':     task_result.status,
        'ready':      task_result.ready(),
        'successful': task_result.successful() if task_result.ready() else None,
        'failed':     task_result.failed()     if task_result.ready() else None,
    }
    if task_result.ready():
        if task_result.successful():
            response['result'] = task_result.result
        elif task_result.failed():
            response['error'] = str(task_result.result)
    return Response(response)


# ============================================================
# STATEMENT UPLOAD & PARSING VIEWS
# ============================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def upload_to_drive(request):
    try:
        bank_name          = request.data.get('bank_name')
        statement_folder_id = request.data.get('statement_folder_id')
        uploaded_file      = request.FILES.get('file')

        if not all([bank_name, statement_folder_id, uploaded_file]):
            return Response(
                {'error': 'bank_name, statement_folder_id, and file are required'},
                status=400
            )

        allowed = {
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/pdf', 'image/jpeg', 'image/png',
        }
        if uploaded_file.content_type not in allowed:
            return Response({'error': f'Unsupported file type: {uploaded_file.content_type}'}, status=400)

        result = drive_service.upload_file(
            file_obj=uploaded_file,
            file_name=uploaded_file.name,
            folder_id=statement_folder_id,
            mime_type=uploaded_file.content_type,
        )
        if not result:
            return Response({'error': 'Failed to upload to Drive'}, status=500)

        return Response({'success': True, 'message': 'File uploaded', 'file': result},
                        status=status.HTTP_201_CREATED)

    except Exception as e:
        logger.error(f"Upload error: {e}")
        return Response({'error': str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def upload_to_drive_async(request):
    bank_name          = request.data.get('bank_name')
    statement_folder_id = request.data.get('statement_folder_id')
    uploaded_file      = request.FILES.get('file')

    if not all([bank_name, statement_folder_id, uploaded_file]):
        return Response({'error': 'bank_name, statement_folder_id, and file are required'}, status=400)

    allowed = {
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/pdf', 'image/jpeg', 'image/png',
    }
    if uploaded_file.content_type not in allowed:
        return Response({'error': f'Unsupported file type: {uploaded_file.content_type}'}, status=400)

    task = upload_to_drive_task.delay(
        uploaded_file.read(), uploaded_file.name,
        statement_folder_id, uploaded_file.content_type,
    )
    return Response({
        'success':    True,
        'message':    'File upload queued',
        'task_id':    task.id,
        'status_url': f"/api/tasks/{task.id}/status/",
    }, status=status.HTTP_202_ACCEPTED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def parse_statement(request):
    bank_account_id = request.data.get('bank_account_id')
    uploaded_file   = request.FILES.get('file')

    if not bank_account_id or not uploaded_file:
        return Response({'error': 'bank_account_id and file are required'}, status=400)

    try:
        bank_account = BankAccount.objects.get(
            id=bank_account_id, company=request.user.company
        )
    except BankAccount.DoesNotExist:
        return Response({'error': 'Bank account not found'}, status=404)

    file_bytes = uploaded_file.read()

    stmt = UploadedStatement.objects.create(
        bank_account=bank_account,
        file_name=uploaded_file.name,
        parsed=False,
    )

    result = drive_service.upload_file(
        file_obj=io.BytesIO(file_bytes),
        file_name=uploaded_file.name,
        folder_id=bank_account.statement_folder_id,
        mime_type=uploaded_file.content_type,
    )

    if not result:
        stmt.delete()
        return Response({'error': 'Failed to upload file to Drive'}, status=500)

    stmt.drive_file_id = result.get('file_id')
    stmt.save(update_fields=['drive_file_id'])

    task = parse_statement_task.delay(stmt.id)

    return Response({
        'success':      True,
        'message':      'Statement uploaded and queued for processing',
        'statement_id': stmt.id,
        'task_id':      task.id,
        'status_url':   f"/api/tasks/{task.id}/status/",
    }, status=status.HTTP_202_ACCEPTED)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_statements(request):
    qs = UploadedStatement.objects.filter(
        bank_account__company=request.user.company
    ).order_by('-uploaded_at')

    bank_account_id = request.query_params.get('bank_account_id')
    if bank_account_id:
        qs = qs.filter(bank_account_id=bank_account_id)

    return Response(UploadedStatementSerializer(qs, many=True).data)


# ============================================================
# RECEIPT MANAGEMENT VIEWS
# ============================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def upload_receipt(request):
    try:
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'error': 'file is required'}, status=400)

        allowed = {
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/pdf', 'image/jpeg', 'image/png',
        }
        if uploaded_file.content_type not in allowed:
            return Response({'error': f'Unsupported file type: {uploaded_file.content_type}'}, status=400)

        company_name = _get_company_name(request)
        if not company_name:
            return Response({'error': 'Could not determine company name'}, status=400)

        company_result = drive_service.get_or_create_company_folder(company_name)
        if not company_result:
            return Response({'error': 'Could not access company Drive folder'}, status=500)

        billing_folder_id = company_result['subfolders'].get('Billing_Receipt')
        if not billing_folder_id:
            return Response({'error': 'Billing_Receipt folder not found'}, status=500)

        file_bytes = uploaded_file.read()

        result = drive_service.upload_file(
            file_obj=io.BytesIO(file_bytes),
            file_name=uploaded_file.name,
            folder_id=billing_folder_id,
            mime_type=uploaded_file.content_type,
        )
        if not result:
            return Response({'error': 'Failed to upload to Drive'}, status=500)

        # FIX: use timezone.now() not datetime.now() to avoid TZ crash
        receipt = ReceiptDocument.objects.create(
            company=request.user.company,
            drive_file_id=result.get('file_id'),
            file_name=uploaded_file.name,
            file_link=result.get('file_link'),
            mime_type=uploaded_file.content_type,
            uploaded_at=timezone.now(),
            extracted=False,
        )

        task = process_receipt_task.delay(receipt.id)

        return Response({
            'success':    True,
            'message':    'Receipt uploaded and queued for processing',
            'receipt_id': receipt.id,
            'task_id':    task.id,
            'status_url': f"/api/tasks/{task.id}/status/",
        }, status=status.HTTP_202_ACCEPTED)

    except Exception as e:
        logger.error(f"Error uploading receipt: {e}", exc_info=True)
        return Response({'error': str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def sync_receipt_folder(request):
    company_name = _get_company_name(request)
    if not company_name:
        return Response({'error': 'Could not determine company name'}, status=400)

    company_result = drive_service.get_or_create_company_folder(company_name)
    if not company_result:
        return Response({'error': 'Could not access company Drive folder'}, status=500)

    billing_folder_id = company_result['subfolders'].get('Billing_Receipt')
    if not billing_folder_id:
        return Response({'error': 'Billing_Receipt folder not found'}, status=500)

    files         = drive_service.list_files(billing_folder_id)
    created_count = 0
    updated_count = 0

    for f in files:
        # FIX: Drive timestamps are ISO format '2026-06-28T10:20:59.000Z'
        # use _parse_drive_timestamp() not _parse_date()
        uploaded_at = _parse_drive_timestamp(f.get('createdTime', ''))

        _, created = ReceiptDocument.objects.update_or_create(
            company=request.user.company,
            drive_file_id=f.get('id', ''),
            defaults={
                'file_name':  f.get('name', ''),
                'file_link':  f.get('webViewLink', ''),
                'mime_type':  f.get('mimeType', ''),
                'uploaded_at': uploaded_at or timezone.now(),
            },
        )
        if created:
            created_count += 1
        else:
            updated_count += 1

    return Response({
        'success':          True,
        'folder_id':        billing_folder_id,
        'total_files_seen': len(files),
        'created':          created_count,
        'updated':          updated_count,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_receipts(request):
    # FIX: order only by uploaded_at — created_at may not exist on the model
    qs = ReceiptDocument.objects.filter(
        company=request.user.company
    ).order_by('-uploaded_at')

    extracted = request.query_params.get('extracted')
    if extracted in ('true', 'false'):
        qs = qs.filter(extracted=(extracted == 'true'))

    search = request.query_params.get('search')
    if search:
        qs = qs.filter(
            Q(file_name__icontains=search) | Q(receipt_no__icontains=search)
        )

    page      = int(request.query_params.get('page', 1))
    page_size = int(request.query_params.get('page_size', 50))
    start     = (page - 1) * page_size
    end       = start + page_size
    total     = qs.count()

    return Response({
        'count':   total,
        'page':    page,
        'pages':   (total + page_size - 1) // page_size,
        'results': ReceiptDocumentSerializer(qs[start:end], many=True).data,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def extract_receipts(request):
    receipt_ids = request.data.get('receipt_ids')
    qs = ReceiptDocument.objects.filter(company=request.user.company, extracted=False)
    if receipt_ids:
        qs = qs.filter(id__in=receipt_ids)

    if not qs.exists():
        return Response({
            'success': True, 'message': 'No receipts to extract',
            'processed': 0, 'extracted': 0, 'failed': 0,
        })

    processed    = 0
    extracted_ok = 0
    failed       = 0
    errors       = []

    for receipt in qs:
        processed += 1
        try:
            file_bytes = drive_service.download_file(receipt.drive_file_id)
            if not file_bytes:
                receipt.extracted        = False
                receipt.extraction_error = 'Failed to download from Drive'
                receipt.save(update_fields=['extracted', 'extraction_error'])
                failed += 1
                errors.append(f"{receipt.file_name}: Download failed")
                continue

            result = parse_receipt_pdf(file_bytes, receipt.file_name)

            receipt.receipt_no   = result.get('receipt_no')   or receipt.receipt_no
            receipt.receipt_date = result.get('receipt_date') or receipt.receipt_date

            amount_val = result.get('amount')
            if amount_val is None:
                receipt.extracted        = False
                receipt.extraction_error = result.get('error', 'Could not extract amount')
                failed += 1
                errors.append(f"{receipt.file_name}: {receipt.extraction_error}")
            else:
                receipt.amount           = amount_val
                receipt.extracted        = True
                receipt.extraction_error = ''
                extracted_ok += 1

            receipt.save(update_fields=[
                'receipt_no', 'receipt_date', 'amount', 'extracted', 'extraction_error'
            ])

        except Exception as exc:
            logger.error(f"Receipt {receipt.file_name}: {exc}", exc_info=True)
            receipt.extracted        = False
            receipt.extraction_error = str(exc)
            receipt.save(update_fields=['extracted', 'extraction_error'])
            failed += 1
            errors.append(f"{receipt.file_name}: {str(exc)}")

    return Response({
        'success':   True,
        'message':   f'Extracted {extracted_ok}, failed {failed}',
        'processed': processed,
        'extracted': extracted_ok,
        'failed':    failed,
        'errors':    errors[:10],
    })


# ============================================================
# TRANSACTION VIEWS
# ============================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_transactions(request):
    qs = Transaction.objects.filter(company=request.user.company)

    bank_account_id = request.query_params.get('bank_account_id')
    rec_status      = request.query_params.get('status')
    date_from       = request.query_params.get('date_from')
    date_to         = request.query_params.get('date_to')
    search          = request.query_params.get('search')

    if bank_account_id: qs = qs.filter(bank_account_id=bank_account_id)
    if rec_status:      qs = qs.filter(reconcile_status=rec_status)
    if date_from:       qs = qs.filter(txn_date__gte=date_from)
    if date_to:         qs = qs.filter(txn_date__lte=date_to)
    if search:
        qs = qs.filter(
            Q(description__icontains=search) |
            Q(utr_no__icontains=search)      |
            Q(ref_no__icontains=search)
        )

    qs        = qs.select_related('bank_account__bank').order_by('-txn_date')
    page      = int(request.query_params.get('page', 1))
    page_size = int(request.query_params.get('page_size', 50))
    start     = (page - 1) * page_size
    total     = qs.count()

    return Response({
        'count':   total,
        'page':    page,
        'pages':   (total + page_size - 1) // page_size,
        'results': TransactionSerializer(qs[start:start + page_size], many=True).data,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reconciliation_stats(request):
    company = request.user.company
    acc1_id = request.query_params.get('bank_account_1')
    acc2_id = request.query_params.get('bank_account_2')

    qs = Transaction.objects.filter(company=company)
    if acc1_id and acc2_id:
        qs = qs.filter(bank_account_id__in=[acc1_id, acc2_id])

    return Response(qs.aggregate(
        total=Count('id'),
        total_matched=Count('id', filter=Q(reconcile_status='matched')),
        total_unmatched=Count('id', filter=Q(reconcile_status='unmatched')),
        total_receipt_missing=Count('id', filter=Q(reconcile_status='receipt_missing')),
    ))


# ============================================================
# RECONCILIATION VIEWS
# ============================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def run_reconciliation_view(request):
    acc1_id   = request.data.get('bank_account_1')
    acc2_id   = request.data.get('bank_account_2')
    date_from = request.data.get('date_from')
    date_to   = request.data.get('date_to')

    if not acc1_id or not acc2_id:
        return Response({'error': 'bank_account_1 and bank_account_2 are required'}, status=400)

    result = run_reconciliation(
        bank_account_1_id=int(acc1_id),
        bank_account_2_id=int(acc2_id),
        company_id=request.user.company.id,
        date_from=date_from,
        date_to=date_to,
        # user=request.user,
    )

    if 'error' in result:
        return Response(result, status=400)
    return Response({'success': True, **result})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_summary(request):
    company   = request.user.company
    cache_key = f'dashboard_summary_{company.id}'
    cached    = cache.get(cache_key)
    if cached:
        return Response(cached)

    bank_accounts = BankAccount.objects.filter(company=company)
    transactions  = Transaction.objects.filter(company=company)

    total_transactions = transactions.count()
    matched_count      = transactions.filter(reconcile_status='matched').count()
    receipt_missing    = transactions.filter(reconcile_status='receipt_missing').count()
    unmatched          = transactions.filter(reconcile_status='unmatched').count()
    ignored            = transactions.filter(reconcile_status='ignored').count()

    # FIX: use separate denominator variable — don't overwrite total_transactions
    denom = total_transactions or 1
    matched_pct        = matched_count   / denom * 100
    receipt_missing_pct = receipt_missing / denom * 100
    unmatched_pct      = unmatched       / denom * 100

    last_reconciliation = ReconciliationRun.objects.filter(company=company).first()

    bank_summaries = []
    for ba in bank_accounts:
        last_stmt = ba.statements.filter(parsed=True).order_by('-uploaded_at').first()
        bank_summaries.append({
            'id':               ba.id,
            'bank_name':        ba.bank.bank_name,
            'account_holder':   ba.account_holder_name,
            'account_number':   ba.account_number,
            'ifsc':             ba.ifsc_code or 'N/A',
            'last_synced':      last_stmt.uploaded_at if last_stmt else None,
            'transaction_count': ba.transactions.count(),
            'matched_count':    ba.transactions.filter(reconcile_status='matched').count(),
        })

    # 24h stats — FIX: keep raw count separate from the display value
    last_24h          = timezone.now() - timezone.timedelta(hours=24)
    txns_24h          = transactions.filter(created_at__gte=last_24h)
    total_24h_raw     = txns_24h.count()
    matched_24h       = txns_24h.filter(reconcile_status='matched').count()
    missing_24h       = txns_24h.filter(reconcile_status='receipt_missing').count()
    unmatched_24h     = txns_24h.filter(reconcile_status='unmatched').count()
    denom_24h         = total_24h_raw or 1   # separate variable — don't corrupt total_24h_raw

    data = {
        'total_transactions':       total_transactions,
        'matched':                  matched_count,
        'matched_percentage':       round(matched_pct, 1),
        'receipt_missing':          receipt_missing,
        'receipt_missing_percentage': round(receipt_missing_pct, 1),
        'unmatched':                unmatched,
        'unmatched_percentage':     round(unmatched_pct, 1),
        'ignored':                  ignored,
        'last_reconciliation':      last_reconciliation.run_date if last_reconciliation else None,
        'last_reconciliation_id':   last_reconciliation.id       if last_reconciliation else None,
        'banks':                    bank_summaries,
        'last_24h': {
            'total':                     total_24h_raw,   # FIX: real count not the `or 1` value
            'matched':                   matched_24h,
            'matched_percentage':        round(matched_24h / denom_24h * 100, 1),
            'receipt_missing':           missing_24h,
            'receipt_missing_percentage': round(missing_24h / denom_24h * 100, 1),
            'unmatched':                 unmatched_24h,
            'unmatched_percentage':      round(unmatched_24h / denom_24h * 100, 1),
        },
    }

    cache.set(cache_key, data, 300)
    return Response(data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def run_full_reconciliation_view(request):
    company    = request.user.company
    acc1_id    = request.data.get('bank_account_1')
    acc2_id    = request.data.get('bank_account_2')
    date_from  = request.data.get('date_from')
    date_to    = request.data.get('date_to')
    auto_match = request.data.get('auto_match', True)
    async_mode = request.data.get('async', False)

    if not acc1_id or not acc2_id:
        return Response({'error': 'bank_account_1 and bank_account_2 are required'}, status=400)

    try:
        if async_mode:
            task = run_reconciliation_task.delay(
                company_id=company.id,
                bank_account_1_id=int(acc1_id),
                bank_account_2_id=int(acc2_id),
                date_from=date_from,
                date_to=date_to,
                auto_match=auto_match,
            )
            return Response({
                'success':    True,
                'message':    'Reconciliation queued',
                'task_id':    task.id,
                'status_url': f"/api/tasks/{task.id}/status/",
            }, status=status.HTTP_202_ACCEPTED)

        result = run_full_reconciliation(
            company=company,
            user=request.user,
            bank_account_1_id=int(acc1_id),
            bank_account_2_id=int(acc2_id),
            date_from=date_from,
            date_to=date_to,
            auto_match=auto_match,
        )

        AuditLog.objects.create(
            company=company,
            user=request.user,
            action='AUTO_MATCH' if auto_match else 'RUN_RECONCILIATION',
            entity_type='RECONCILIATION',
            entity_id=result['run_id'],
            details={
                'bank_account_1': acc1_id,
                'bank_account_2': acc2_id,
                'summary': result['summary'],
            },
        )
        cache.delete(f'dashboard_summary_{company.id}')
        return Response({'success': True, **result})

    except Exception as e:
        logger.error(f"Reconciliation failed: {e}", exc_info=True)
        return Response({'error': str(e)}, status=500)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def run_full_reconciliation_v2(request):

    company = request.user.company

    bank_account_1 = request.data.get("bank_account_1")
    bank_account_2 = request.data.get("bank_account_2")

    date_from = request.data.get("date_from")
    date_to = request.data.get("date_to")

    auto_match = request.data.get("auto_match", True)
    async_mode = request.data.get("async", False)

    if not bank_account_1 or not bank_account_2:
        return Response(
            {
                "success": False,
                "error": "bank_account_1 and bank_account_2 are required",
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:

        if async_mode:

            task = run_reconciliation_task.delay(
                company_id=company.id,
                bank_account_1_id=int(bank_account_1),
                bank_account_2_id=int(bank_account_2),
                date_from=date_from,
                date_to=date_to,
                auto_match=auto_match,
            )

            return Response(
                {
                    "success": True,
                    "task_id": task.id,
                    "status_url": f"/api/tasks/{task.id}/status/",
                },
                status=status.HTTP_202_ACCEPTED,
            )

        result = run_reconciliation_v2(
            company=company,
            user=request.user,
            bank_account_1_id=int(bank_account_1),
            bank_account_2_id=int(bank_account_2),
            date_from=date_from,
            date_to=date_to,
            auto_match=auto_match,
        )

        AuditLog.objects.create(
            company=company,
            user=request.user,
            action="RUN_RECONCILIATION",
            entity_type="RECONCILIATION",
            entity_id=result["run_id"],
            details=result["summary"],
        )

        cache.delete(f"dashboard_summary_{company.id}")

        return Response(
            {
                "success": True,
                **result,
            }
        )

    except Exception as e:

        logger.exception(e)

        return Response(
            {
                "success": False,
                "error": str(e),
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def match_receipts_view(request):
    company        = request.user.company
    bank_account_id = request.data.get('bank_account_id')

    result = match_with_receipts_only(
        company=company,
        bank_account_id=bank_account_id,
        user=request.user,
    )

    AuditLog.objects.create(
        company=company,
        user=request.user,
        action='AUTO_MATCH',
        entity_type='RECEIPT_MATCHING',
        entity_id=0,
        details={'result': result},
    )
    cache.delete(f'dashboard_summary_{company.id}')
    return Response({'success': True, **result})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reconciliation_history(request):
    company = request.user.company
    runs    = ReconciliationRun.objects.filter(company=company)

    bank_id   = request.query_params.get('bank_id')
    date_from = request.query_params.get('date_from')
    date_to   = request.query_params.get('date_to')

    if bank_id:
        runs = runs.filter(Q(bank_account_1_id=bank_id) | Q(bank_account_2_id=bank_id))
    if date_from: runs = runs.filter(run_date__gte=date_from)
    if date_to:   runs = runs.filter(run_date__lte=date_to)

    page      = int(request.query_params.get('page', 1))
    page_size = int(request.query_params.get('page_size', 20))
    start     = (page - 1) * page_size
    total     = runs.count()

    return Response({
        'count':   total,
        'page':    page,
        'pages':   (total + page_size - 1) // page_size,
        'results': ReconciliationRunSerializer(runs[start:start + page_size], many=True).data,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_audit_log(request):
    company    = request.user.company
    audit_logs = AuditLog.objects.filter(company=company)

    action    = request.query_params.get('action')
    date_from = request.query_params.get('date_from')
    date_to   = request.query_params.get('date_to')
    user_id   = request.query_params.get('user_id')

    if action:    audit_logs = audit_logs.filter(action=action)
    if date_from: audit_logs = audit_logs.filter(timestamp__gte=date_from)
    if date_to:   audit_logs = audit_logs.filter(timestamp__lte=date_to)
    if user_id:   audit_logs = audit_logs.filter(user_id=user_id)

    page      = int(request.query_params.get('page', 1))
    page_size = int(request.query_params.get('page_size', 50))
    start     = (page - 1) * page_size
    total     = audit_logs.count()

    return Response({
        'count':   total,
        'page':    page,
        'pages':   (total + page_size - 1) // page_size,
        'results': AuditLogSerializer(audit_logs[start:start + page_size], many=True).data,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def clear_all_data(request):
    company = request.user.company

    if not request.data.get('confirm', False):
        return Response({'error': 'Confirmation required. Set confirm=true'}, status=400)

    txn_count  = Transaction.objects.filter(company=company).count()
    stmt_count = UploadedStatement.objects.filter(bank_account__company=company).count()
    run_count  = ReconciliationRun.objects.filter(company=company).count()

    Transaction.objects.filter(company=company).delete()
    UploadedStatement.objects.filter(bank_account__company=company).delete()
    ReconciliationRun.objects.filter(company=company).delete()

    AuditLog.objects.create(
        company=company,
        user=request.user,
        action='CLEAR_ALL_DATA',
        entity_type='ALL',
        entity_id=0,
        details={
            'transactions_deleted': txn_count,
            'statements_deleted':   stmt_count,
            'runs_deleted':         run_count,
        },
    )
    cache.delete(f'dashboard_summary_{company.id}')

    return Response({
        'message': 'All data cleared successfully',
        'deleted': {
            'transactions':        txn_count,
            'statements':          stmt_count,
            'reconciliation_runs': run_count,
        },
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def extraction_status(request):
    company      = request.user.company
    statements   = UploadedStatement.objects.filter(bank_account__company=company)
    transactions = Transaction.objects.filter(company=company)
    receipts     = ReceiptDocument.objects.filter(company=company)

    total            = transactions.count()
    matched          = transactions.filter(reconcile_status='matched').count()
    receipt_missing  = transactions.filter(reconcile_status='receipt_missing').count()
    unmatched        = transactions.filter(reconcile_status='unmatched').count()
    total_receipts   = receipts.count()
    extracted        = receipts.filter(extracted=True).count()

    return Response({
        'statements': {
            'total':  statements.count(),
            'parsed': statements.filter(parsed=True).count(),
            'failed': statements.filter(parsed=False).count(),
        },
        'transactions': {
            'total':          total,
            'matched':        matched,
            'receipt_missing': receipt_missing,
            'unmatched':      unmatched,
            'match_percentage': round(matched / total * 100 if total else 0, 1),
        },
        'receipts': {
            'total':                total_receipts,
            'extracted':            extracted,
            'failed':               total_receipts - extracted,
            'extraction_percentage': round(extracted / total_receipts * 100 if total_receipts else 0, 1),
        },
        'is_complete': (
            statements.filter(parsed=False).count() == 0 and
            receipts.filter(extracted=False).count() == 0 and
            total > 0
        ),
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def verification_extraction(request):
    company      = request.user.company
    statements   = UploadedStatement.objects.filter(bank_account__company=company)
    transactions = Transaction.objects.filter(company=company)
    receipts     = ReceiptDocument.objects.filter(company=company)

    total_receipts     = receipts.count()
    extracted_receipts = receipts.filter(extracted=True).count()
    matched            = transactions.filter(reconcile_status='matched').count()
    receipt_missing    = transactions.filter(reconcile_status='receipt_missing').count()
    unmatched          = transactions.filter(reconcile_status='unmatched').count()

    failed_stmts = [
        {'file_name': s.file_name, 'error': s.parse_error or 'Unknown'}
        for s in statements.filter(parsed=False)
    ]
    failed_receipts = [
        {'file_name': r.file_name, 'error': r.extraction_error or 'Unknown'}
        for r in receipts.filter(extracted=False)
    ]

    return Response({
        'success': True,
        'summary': {
            'total_statements':    statements.count(),
            'parsed_statements':   statements.filter(parsed=True).count(),
            'failed_statements':   statements.filter(parsed=False).count(),
            'total_transactions':  transactions.count(),
            'total_receipts':      total_receipts,
            'extracted_receipts':  extracted_receipts,
            'receipts_with_amount': receipts.filter(extracted=True, amount__isnull=False).count(),
            'receipts_with_date':  receipts.filter(extracted=True, receipt_date__isnull=False).count(),
        },
        'reconciliation_status': {
            'matched':         matched,
            'receipt_missing': receipt_missing,
            'unmatched':       unmatched,
        },
        'failed_statements': failed_stmts,
        'failed_receipts':   failed_receipts,
        'is_complete': (
            statements.filter(parsed=False).count() == 0 and
            not failed_receipts and
            transactions.count() > 0
        ),
    })

from rest_framework.pagination import PageNumberPagination
class TransactionPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100

from .serializers import TransactionManagementSerializer

# @api_view(["GET"])
# @permission_classes([IsAuthenticated])
# def transaction_management(request):
#     company = request.user.company

#     queryset = (
#         Transaction.objects
#         .filter(company=company)
#         .select_related(
#             "bank_account",
#             "bank_account__bank",
#             "matched_txn",
#             "matched_txn__bank_account",
#             "matched_txn__bank_account__bank",
#             "particular",
#             "accounting_particular",
#         )
#         .prefetch_related("matched_receipts")
#     )
#     bank_account_id = request.GET.get("bank_account_id")
    
#     if bank_account_id:
#         queryset = queryset.filter(bank_account_id=bank_account_id)

#     status = request.GET.get("status")
    
#     if status:
#         queryset = queryset.filter(reconcile_status=status)

#     transaction_type = request.GET.get("transaction_type")

#     if transaction_type:

#         if transaction_type == "credit":

#             queryset = queryset.filter(
#                 txn_type="credit"
#             )

#         elif transaction_type == "debit":

#             queryset = queryset.filter(
#                 txn_type="debit"
#             )

#         elif transaction_type == "internal_transfer":

#             queryset = queryset.filter(
#                 reconcile_status="ignored",
#                 matched_txn__isnull=False
#             )

#     date_from = request.GET.get("date_from")

#     date_to = request.GET.get("date_to")

#     if date_from:

#         queryset = queryset.filter(
#             txn_date__gte=date_from
#         )

#     if date_to:

#         queryset = queryset.filter(
#             txn_date__lte=date_to
#         )

#     search = request.GET.get("search")

#     if search:

#         queryset = queryset.filter(

#             Q(description__icontains=search)

#             |

#             Q(ref_no__icontains=search)

#             |

#             Q(utr_no__icontains=search)

#         )

#     ordering = request.GET.get("ordering", "-txn_date")

#     allowed_ordering = [
#         "txn_date",
#         "-txn_date",
#         "amount",
#         "-amount",
#         "created_at",
#         "-created_at",
#     ]

#     if ordering not in allowed_ordering:
#         ordering = "-txn_date"

#     if ordering == "txn_date":
#         queryset = queryset.order_by("txn_date", "id")

#     elif ordering == "-txn_date":
#         queryset = queryset.order_by("-txn_date", "-id")

#     elif ordering == "amount":
#         queryset = queryset.order_by("amount", "id")

#     elif ordering == "-amount":
#         queryset = queryset.order_by("-amount", "-id")

#     elif ordering == "created_at":
#         queryset = queryset.order_by("created_at", "id")

#     elif ordering == "-created_at":
#         queryset = queryset.order_by("-created_at", "-id")

#     paginator = TransactionPagination()

#     page = paginator.paginate_queryset(
#         queryset,
#         request
#     )

#     serializer = TransactionManagementSerializer(page, many=True)
#     # serializer = TransactionManagementSerializer(queryset, many=True)
#     return paginator.get_paginated_response(
#         {
#             "success": True,
#             "message": "Transactions fetched successfully.",
#             "results": serializer.data,
#         }
#     )
#     # return Response(
#     #     {
#     #         "success": True,
#     #         "message": "Transactions fetched successfully.",
#     #         "results": serializer.data,
#     #     }
#     # )



@api_view(["GET"])
@permission_classes([IsAuthenticated])
def transaction_management(request):

    company = request.user.company

    queryset = (
        Transaction.objects
        .filter(company=company)
        .select_related(
            "bank_account",
            "bank_account__bank",
            "matched_txn",
            "matched_txn__bank_account",
            "matched_txn__bank_account__bank",
        )
    )

    # =====================================================
    # Bank Account Filter
    # =====================================================

    bank_account_id = request.GET.get(
        "bank_account_id"
    )

    if bank_account_id:

        queryset = queryset.filter(
            bank_account_id=bank_account_id
        )

    # =====================================================
    # Status Filter
    # =====================================================

    status_value = request.GET.get(
        "status"
    )

    if status_value:

        queryset = queryset.filter(
            reconcile_status=status_value
        )

    # =====================================================
    # Transaction Type Filter
    # =====================================================

    transaction_type = request.GET.get(
        "transaction_type"
    )

    if transaction_type == "credit":

        queryset = queryset.filter(
            txn_type="credit"
        )

    elif transaction_type == "debit":

        queryset = queryset.filter(
            txn_type="debit"
        )

    elif transaction_type == "internal_transfer":

        queryset = queryset.filter(
            reconcile_status="ignored",
            matched_txn__isnull=False
        )

    # =====================================================
    # Date Filters
    # =====================================================

    date_from = request.GET.get(
        "date_from"
    )

    date_to = request.GET.get(
        "date_to"
    )

    if date_from:

        queryset = queryset.filter(
            txn_date__gte=date_from
        )

    if date_to:

        queryset = queryset.filter(
            txn_date__lte=date_to
        )

    # =====================================================
    # Search
    # =====================================================

    search = request.GET.get(
        "search"
    )

    if search:

        queryset = queryset.filter(

            Q(description__icontains=search)

            |

            Q(ref_no__icontains=search)

            |

            Q(utr_no__icontains=search)
        )

    # =====================================================
    # Ordering
    # =====================================================

    ordering = request.GET.get(
        "ordering",
        "-txn_date"
    )

    allowed_ordering = [
        "txn_date",
        "-txn_date",
        "amount",
        "-amount",
        "created_at",
        "-created_at",
    ]

    if ordering not in allowed_ordering:

        ordering = "-txn_date"

    if ordering == "txn_date":

        queryset = queryset.order_by(
            "txn_date",
            "id"
        )

    elif ordering == "-txn_date":

        queryset = queryset.order_by(
            "-txn_date",
            "-id"
        )

    elif ordering == "amount":

        queryset = queryset.order_by(
            "amount",
            "id"
        )

    elif ordering == "-amount":

        queryset = queryset.order_by(
            "-amount",
            "-id"
        )

    elif ordering == "created_at":

        queryset = queryset.order_by(
            "created_at",
            "id"
        )

    elif ordering == "-created_at":

        queryset = queryset.order_by(
            "-created_at",
            "-id"
        )

    # =====================================================
    # Pagination
    # =====================================================

    paginator = TransactionPagination()

    page = paginator.paginate_queryset(
        queryset,
        request
    )

    serializer = TransactionManagementSerializer(
        page,
        many=True
    )

    return paginator.get_paginated_response(
        {
            "success": True,
            "message": (
                "Transactions fetched successfully."
            ),
            "results": serializer.data,
        }
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def particular_list(request):

    queryset = Particular.objects.filter(
        is_active=True
    )

    transaction_type = request.GET.get("transaction_type")

    if transaction_type:

        queryset = queryset.filter(
            transaction_type=transaction_type.lower()
        )

    serializer = ParticularSerializer(
        queryset.order_by("name"),
        many=True
    )

    return Response({
        "success": True,
        "results": serializer.data
    })




@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def update_transaction(request, pk):
    print(request.data)

    company = request.user.company

    transaction = get_object_or_404(
        Transaction,
        pk=pk,
        company=company
    )

    serializer = TransactionUpdateSerializer(
        transaction,
        data=request.data,
        partial=True
    )

    if serializer.is_valid():

        serializer.save()

        return Response(
            {
                "success": True,
                "message": "Transaction updated successfully.",
                "data": {
                    "id": transaction.id,
                    "particular": {
                        "id": transaction.particular.id if transaction.particular else None,
                        "name": transaction.particular.name if transaction.particular else None,
                    },
                    "accounting_particular": {
                    "id": transaction.accounting_particular.id if transaction.accounting_particular else None,
                    "name": transaction.accounting_particular.name if transaction.accounting_particular else None,
                }
                }
            },
            status=status.HTTP_200_OK
        )

    return Response(
        {
            "success": False,
            "errors": serializer.errors,
        },
        status=status.HTTP_400_BAD_REQUEST
    )




@api_view(["GET"])
@permission_classes([IsAuthenticated])
def cash_flow_analysis(request):
    try:

        service = CashFlowAnalysisService(

            company=request.user.company,

            month=request.GET.get("month"),

            year=request.GET.get("year"),

            bank_account=request.GET.get("bank_account"),

        )

        return Response(
            service.execute()
        )

    except ValueError as e:

        return Response(
            {
                "success": False,
                "message": str(e),
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    except Exception as e:

        return Response(
            {
                "success": False,
                "message": str(e),
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )



@api_view(["GET"])
@permission_classes([IsAuthenticated])
def accounting_particular_list(request):
    queryset = AccountingParticular.objects.filter(
        is_active=True
    )

    serializer = AccountingParticularDropdownSerializer(
        queryset.order_by("name"),
        many=True
    )

    return Response(
        {
            "success": True,
            "results": serializer.data
        },
        status=status.HTTP_200_OK
    )





@api_view(["GET"])
@permission_classes([IsAuthenticated])
def transaction_journey(request, pk):

    company = request.user.company

    transaction = get_object_or_404(

        Transaction.objects.select_related(

            "bank_account",

            "bank_account__bank",

            "particular",

            "accounting_particular",

            "matched_txn",

        ).prefetch_related(

            "matched_receipts"

        ),

        id=pk,

        company=company,

    )

    serializer = TransactionJourneySerializer(transaction)

    return Response(

        {

            "success": True,

            "message": "Transaction journey fetched successfully.",

            "data": serializer.data,

        },

        status=status.HTTP_200_OK,

    )



@api_view(["POST"])
@permission_classes([IsAuthenticated])
def auto_receipt_matching_api(request):

    try:

        result = auto_receipt_matching(
            company=request.user.company,
            user=request.user,
        )

        return Response(
            {
                "success": True,
                **result,
            },
            status=status.HTTP_200_OK,
        )

    except Exception as e:

        logger.exception(e)

        return Response(
            {
                "success": False,
                "error": str(e),
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )




@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_agency(request):

    agencies = Agency.objects.filter(
        is_active=True
    ).order_by('name')

    serializer = ListAgencySerializer(
        agencies,
        many=True
    )

    return Response(
        {
            'success': True,
            'agencies': serializer.data
        },
        status=status.HTTP_200_OK
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_agency(request):

    serializer = CreateAgencySerializer(
        data=request.data
    )

    if not serializer.is_valid():

        return Response(
            {
                'success': False,
                'errors': serializer.errors
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    name = serializer.validated_data.get('name').strip()

    # ---------------------------------------------
    # Duplicate agency validation
    # ---------------------------------------------

    if Agency.objects.filter(
        name__iexact=name
    ).exists():

        return Response(
            {
                'success': False,
                'error': 'Agency with this name already exists.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    try:

        agency = Agency.objects.create(
            **serializer.validated_data
        )

        return Response(
            {
                'success': True,
                'message': 'Agency created successfully.',
                'agency': CreateAgencySerializer(
                    agency
                ).data
            },
            status=status.HTTP_201_CREATED
        )

    except Exception as e:

        logger.error(
            f'Error creating agency: {e}',
            exc_info=True
        )

        return Response(
            {
                'success': False,
                'error': 'Failed to create agency.'
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def agency_detail(request, agency_id):

#     try:
#         grant_master = Agency.objects.get(
#             id=agency_id,
#             is_active=True
#         )

#     except GrantMaster.DoesNotExist:
#         return Response(
#             {
#                 'success': False,
#                 'error': 'Grant not found'
#             },
#             status=status.HTTP_404_NOT_FOUND
#         )

#     return Response(
#         {
#             'success': True,
#             'grant_master': GrantMasterSerializer(
#                 grant_master
#             ).data
#         },
#         status=status.HTTP_200_OK
#     )





# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def create_grant(request):

#     serializer = GrantSerializer(
#         data=request.data
#     )

#     if not serializer.is_valid():

#         return Response(
#             {
#                 'success': False,
#                 'errors': serializer.errors
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     company = request.user.company

#     if not company:

#         return Response(
#             {
#                 'success': False,
#                 'error': 'User is not associated with a company.'
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     data = serializer.validated_data

#     grant_master_id = data.pop(
#         'grant_master_id',
#         None
#     )

#     grant_name = data.get('name', '').strip()

#     try:

#         with transaction.atomic():

#             # =====================================
#             # EXISTING GRANT MASTER
#             # =====================================

#             if grant_master_id:

#                 grant_master = GrantMaster.objects.get(
#                     id=grant_master_id,
#                     is_active=True
#                 )

#             # =====================================
#             # NEW GRANT MASTER
#             # =====================================

#             else:

#                 grant_master = GrantMaster.objects.filter(
#                     grant_name__iexact=grant_name
#                 ).first()

#                 if not grant_master:

#                     grant_master = GrantMaster.objects.create(
#                         grant_name=grant_name
#                     )

#             # =====================================
#             # CREATE COMPANY GRANT
#             # =====================================

#             grant = Grant.objects.create(
#                 company=company,
#                 grant_master=grant_master,
#                 **data
#             )

#         return Response(
#             {
#                 'success': True,
#                 'message': 'Grant created successfully',
#                 'grant': GrantSerializer(grant).data
#             },
#             status=status.HTTP_201_CREATED
#         )

#     except Exception as e:

#         return Response(
#             {
#                 'success': False,
#                 'error': str(e)
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )




# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def create_grant(request):

#     # -----------------------------------------
#     # 1. Get company from logged-in user
#     # -----------------------------------------

#     company = request.user.company

#     if not company:
#         return Response(
#             {
#                 'success': False,
#                 'error': 'User is not associated with a company.'
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     # -----------------------------------------
#     # 2. Validate request data
#     # -----------------------------------------

#     serializer = GrantSerializer(
#         data=request.data
#     )

#     if not serializer.is_valid():
#         return Response(
#             {
#                 'success': False,
#                 'errors': serializer.errors
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     data = serializer.validated_data

#     # -----------------------------------------
#     # 3. Get selected Agency
#     # -----------------------------------------

#     agency = data.get('agency')

#     if not agency:
#         return Response(
#             {
#                 'success': False,
#                 'error': 'agency_id is required.'
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     # -----------------------------------------
#     # 4. Validate Grant name
#     # -----------------------------------------

#     grant_name = data.get('name', '').strip()

#     if not grant_name:
#         return Response(
#             {
#                 'success': False,
#                 'error': 'Grant name is required.'
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     # -----------------------------------------
#     # 5. Check duplicate Grant
#     # -----------------------------------------

#     existing_grant = Grant.objects.filter(
#         company=company,
#         name__iexact=grant_name
#     ).exists()

#     if existing_grant:
#         return Response(
#             {
#                 'success': False,
#                 'error': 'A grant with this name already exists.'
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     # -----------------------------------------
#     # 6. Create Grant
#     # -----------------------------------------

#     try:

#         with transaction.atomic():

#             grant = Grant.objects.create(
#                 company=company,
#                 agency=agency,
#                 name=grant_name,
#                 amount=data.get('amount'),
#                 start_date=data.get('start_date'),
#                 end_date=data.get('end_date'),
#                 description=data.get('description')
#             )

#         return Response(
#             {
#                 'success': True,
#                 'message': 'Grant created successfully.',
#                 'grant': GrantSerializer(grant).data
#             },
#             status=status.HTTP_201_CREATED
#         )

#     except Exception as e:

#         logger.error(
#             f'Error creating grant: {e}',
#             exc_info=True
#         )

#         return Response(
#             {
#                 'success': False,
#                 'error': str(e)
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_grant(request):

    # =====================================================
    # 1. GET COMPANY FROM LOGGED-IN USER
    # =====================================================

    company = request.user.company

    if not company:
        return Response(
            {
                'success': False,
                'error': 'User is not associated with a company.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # =====================================================
    # 2. GET REQUEST DATA
    # =====================================================

    grant_data = request.data.get('grant')
    transaction_data = request.data.get('transaction')
    milestones_data = request.data.get('milestones')

    if not grant_data:
        return Response(
            {
                'success': False,
                'error': 'Grant information is required.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    if not transaction_data:
        return Response(
            {
                'success': False,
                'error': 'Grant transaction information is required.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    if milestones_data is None:
        return Response(
            {
                'success': False,
                'error': 'Milestones information is required.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # =====================================================
    # 3. VALIDATE GRANT
    # =====================================================

    grant_serializer = GrantSerializer(
        data=grant_data
    )

    if not grant_serializer.is_valid():

        return Response(
            {
                'success': False,
                'section': 'grant',
                'errors': grant_serializer.errors
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    grant_validated_data = grant_serializer.validated_data

    # =====================================================
    # 4. DUPLICATE GRANT CHECK
    # =====================================================

    grant_name = grant_validated_data.get('name', '').strip()

    if not grant_name:

        return Response(
            {
                'success': False,
                'error': 'Grant name is required.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    if Grant.objects.filter(
        company=company,
        name__iexact=grant_name
    ).exists():

        return Response(
            {
                'success': False,
                'error': 'A grant with this name already exists.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # =====================================================
    # 5. VALIDATE GRANT TRANSACTION
    # =====================================================

    grant_transaction_serializer = GrantTransactionSerializer(
        data=transaction_data
    )

    if not grant_transaction_serializer.is_valid():

        return Response(
            {
                'success': False,
                'section': 'transaction',
                'errors': grant_transaction_serializer.errors
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    transaction_validated_data = (
        grant_transaction_serializer.validated_data
    )

    # =====================================================
    # 6. VALIDATE MILESTONES
    # =====================================================

    if not isinstance(milestones_data, list):

        return Response(
            {
                'success': False,
                'section': 'milestones',
                'error': 'Milestones must be an array.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    milestone_serializers = []

    for index, milestone_data in enumerate(milestones_data):

        serializer = GrantMilestoneSerializer(
            data=milestone_data
        )

        if not serializer.is_valid():

            return Response(
                {
                    'success': False,
                    'section': 'milestones',
                    'milestone_index': index,
                    'errors': serializer.errors
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        milestone_serializers.append(serializer)

    # =====================================================
    # 7. VALIDATE TOTAL MILESTONE BUDGET
    # =====================================================

    grant_amount = grant_validated_data.get('amount')

    total_milestone_budget = sum(
        serializer.validated_data.get('budget')
        for serializer in milestone_serializers
    )

    if total_milestone_budget > grant_amount:

        return Response(
            {
                'success': False,
                'section': 'milestones',
                'error': (
                    'Total milestone budget cannot be greater '
                    'than the grant amount.'
                ),
                'grant_amount': str(grant_amount),
                'total_milestone_budget': str(
                    total_milestone_budget
                )
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # =====================================================
    # 8. VALIDATE MILESTONE DATES AGAINST GRANT
    # =====================================================

    grant_start_date = grant_validated_data.get('start_date')
    grant_end_date = grant_validated_data.get('end_date')

    for index, serializer in enumerate(milestone_serializers):

        milestone = serializer.validated_data

        milestone_start = milestone.get('start_date')
        milestone_end = milestone.get('end_date')

        if (
            grant_start_date
            and milestone_start < grant_start_date
        ):

            return Response(
                {
                    'success': False,
                    'section': 'milestones',
                    'milestone_index': index,
                    'error': (
                        'Milestone start date cannot be '
                        'earlier than grant start date.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        if (
            grant_end_date
            and milestone_end > grant_end_date
        ):

            return Response(
                {
                    'success': False,
                    'section': 'milestones',
                    'milestone_index': index,
                    'error': (
                        'Milestone end date cannot be '
                        'later than grant end date.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

    # =====================================================
    # 9. CREATE EVERYTHING ATOMICALLY
    # =====================================================

    try:

        with transaction.atomic():

            # ---------------------------------------------
            # CREATE GRANT
            # ---------------------------------------------

            grant = Grant.objects.create(
                company=company,
                **grant_validated_data
            )

            # ---------------------------------------------
            # CREATE GRANT TRANSACTION
            # ---------------------------------------------

            grant_transaction = GrantTransaction.objects.create(
                grant=grant,
                **transaction_validated_data
            )

            # ---------------------------------------------
            # CREATE MILESTONES
            # ---------------------------------------------

            created_milestones = []

            for serializer in milestone_serializers:

                milestone = GrantMilestone.objects.create(
                    grant=grant,
                    **serializer.validated_data
                )

                created_milestones.append(milestone)

        # =================================================
        # 10. RESPONSE
        # =================================================

        return Response(
            {
                'success': True,
                'message': 'Grant created successfully.',

                'grant': GrantSerializer(
                    grant
                ).data,

                'transaction': GrantTransactionSerializer(
                    grant_transaction
                ).data,

                'milestones': GrantMilestoneSerializer(
                    created_milestones,
                    many=True
                ).data,

                'summary': {
                    'grant_amount': str(grant.amount),

                    'total_milestone_budget': str(
                        total_milestone_budget
                    ),

                    'unallocated_grant_balance': str(
                        grant.amount - total_milestone_budget
                    )
                }
            },
            status=status.HTTP_201_CREATED
        )

    except Exception as e:

        logger.error(
            f'Error creating grant: {e}',
            exc_info=True
        )

        return Response(
            {
                'success': False,
                'error': 'Failed to create grant.'
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_agencies_with_grants(request):

    company = request.user.company

    # ---------------------------------------------
    # 1. Validate company
    # ---------------------------------------------

    if not company:
        return Response(
            {
                'success': False,
                'error': 'User is not associated with a company.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # ---------------------------------------------
    # 2. Get Agencies and only available Grants
    # ---------------------------------------------

    agencies = (
        Agency.objects
        .filter(
            is_active=True,
            grants__company=company,
            grants__is_active=True
        )
        .prefetch_related(
            Prefetch(
                'grants',
                queryset=Grant.objects.filter(
                    company=company,
                    is_active=True,
                    transaction_allocations__isnull=True
                ).order_by('name')
            )
        )
        .distinct()
        .order_by('name')
    )

    # ---------------------------------------------
    # 3. Build response
    # ---------------------------------------------

    agency_data = []

    for agency in agencies:

        grants = agency.grants.all()

        # Don't show an Agency if it has
        # no available Grants
        if not grants:
            continue

        agency_data.append(
            {
                'agency_id': agency.id,
                'agency_name': agency.name,
                'description': agency.description,
                'grants': [
                    {
                        'grant_id': grant.id,
                        'grant_name': grant.name,
                        'amount':grant.amount
                    }
                    for grant in grants
                ]
            }
        )

    # ---------------------------------------------
    # 4. Response
    # ---------------------------------------------

    return Response(
        {
            'success': True,
            'agencies': agency_data
        },
        status=status.HTTP_200_OK
    )


# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def assign_credit_to_grant(request, transaction_id):

#     try:
#         bank_transaction = Transaction.objects.get(
#             id=transaction_id,
#             company=request.user.company
#         )
#     except Transaction.DoesNotExist:
#         return Response(
#             {
#                 'success': False,
#                 'error': 'Transaction not found'
#             },
#             status=status.HTTP_404_NOT_FOUND
#         )

#     fund_name = request.data.get('fund_name')

#     if not fund_name:
#         return Response(
#             {
#                 'success': False,
#                 'error': 'fund_name is required'
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     fund_name = fund_name.strip()

#     if not fund_name:
#         return Response(
#             {
#                 'success': False,
#                 'error': 'fund_name cannot be empty'
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     # ---------------------------------
#     # Credit transaction only
#     # ---------------------------------
#     if bank_transaction.txn_type.lower() != 'credit':
#         return Response(
#             {
#                 'success': False,
#                 'error': 'Only credit transactions can be assigned to a Grant'
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     # ---------------------------------
#     # Grant
#     # ---------------------------------
#     grant_id = request.data.get('grant_id')

#     if not grant_id:
#         return Response(
#             {
#                 'success': False,
#                 'error': 'grant_id is required'
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     try:
#         grant = Grant.objects.get(
#             id=grant_id,
#             company=request.user.company,
#             is_active=True
#         )
#     except Grant.DoesNotExist:
#         return Response(
#             {
#                 'success': False,
#                 'error': 'Grant not found or inactive'
#             },
#             status=status.HTTP_404_NOT_FOUND
#         )

#     # ---------------------------------
#     # Prevent duplicate classification
#     # ---------------------------------
#     if Fund.objects.filter(
#         source_transaction=bank_transaction
#     ).exists():
#         return Response(
#             {
#                 'success': False,
#                 'error': 'This credit transaction has already been assigned'
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     # ---------------------------------
#     # Create Fund + Allocation together
#     # ---------------------------------
#     try:
#         with transaction.atomic():

#             fund = Fund.objects.create(
#             grant=grant,
#             source_transaction=bank_transaction,
#             name=fund_name,
#             amount=bank_transaction.amount
#         )

#             allocation = TransactionFundAllocation.objects.create(
#                 transaction=bank_transaction,
#                 fund=fund,
#                 amount=bank_transaction.amount,
#                 allocation_type='source',
#                 created_by=request.user
#             )

#         return Response(
#             {
#                 'success': True,
#                 'message': 'Credit transaction assigned successfully',
#                 'transaction_id': bank_transaction.id,
#                 'grant_id': grant.id,
#                 'fund': FundSerializer(fund).data,
#                 'allocation_id': allocation.id,
#             },
#             status=status.HTTP_201_CREATED
#         )

#     except Exception as e:

#         return Response(
#             {
#                 'success': False,
#                 'error': str(e)
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def assign_credit_to_grant(request, transaction_id):

    company = request.user.company

    # -------------------------------------------------
    # 1. Validate company
    # -------------------------------------------------

    if not company:
        return Response(
            {
                'success': False,
                'error': 'User is not associated with a company.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # -------------------------------------------------
    # 2. Get bank transaction
    # -------------------------------------------------

    try:
        bank_transaction = Transaction.objects.get(
            id=transaction_id,
            company=company
        )

    except Transaction.DoesNotExist:
        return Response(
            {
                'success': False,
                'error': 'Transaction not found.'
            },
            status=status.HTTP_404_NOT_FOUND
        )

    # -------------------------------------------------
    # 3. Only Credit transactions
    # -------------------------------------------------

    if bank_transaction.txn_type.lower() != 'credit':
        return Response(
            {
                'success': False,
                'error': (
                    'Only credit transactions can be '
                    'assigned to a Grant.'
                )
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # -------------------------------------------------
    # 4. Validate transaction amount
    # -------------------------------------------------

    if bank_transaction.amount <= Decimal('0.00'):
        return Response(
            {
                'success': False,
                'error': (
                    'Credit transaction amount must '
                    'be greater than zero.'
                )
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # -------------------------------------------------
    # 5. Get grant_id
    # -------------------------------------------------

    grant_id = request.data.get('grant_id')

    if not grant_id:
        return Response(
            {
                'success': False,
                'error': 'grant_id is required.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # -------------------------------------------------
    # 6. Validate Grant
    # -------------------------------------------------

    try:
        grant = Grant.objects.get(
            id=grant_id,
            company=company,
            is_active=True
        )

    except Grant.DoesNotExist:
        return Response(
            {
                'success': False,
                'error': 'Grant not found or inactive.'
            },
            status=status.HTTP_404_NOT_FOUND
        )

    # -------------------------------------------------
    # 7. Check whether Grant is already assigned
    # -------------------------------------------------

    if TransactionGrantAllocation.objects.filter(
        grant=grant,
        allocation_type='source'
    ).exists():

        return Response(
            {
                'success': False,
                'error': (
                    'This Grant has already been assigned '
                    'to a credit transaction.'
                ),
                'grant_id': grant.id
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # -------------------------------------------------
    # 8. Check whether this transaction is already assigned
    # -------------------------------------------------

    if TransactionGrantAllocation.objects.filter(
        transaction=bank_transaction,
        allocation_type='source'
    ).exists():

        return Response(
            {
                'success': False,
                'error': (
                    'This credit transaction has already '
                    'been assigned to a Grant.'
                ),
                'transaction_id': bank_transaction.id
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # -------------------------------------------------
    # 9. Validate Grant amount against bank transaction
    # -------------------------------------------------

    if grant.amount != bank_transaction.amount:

        return Response(
            {
                'success': False,
                'error': (
                    'Grant amount does not match the '
                    'credit transaction amount.'
                ),
                'grant_amount': str(grant.amount),
                'transaction_amount': str(
                    bank_transaction.amount
                )
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # -------------------------------------------------
    # 10. Create allocation
    # -------------------------------------------------

    try:

        with transaction.atomic():

            allocation = TransactionGrantAllocation.objects.create(
                transaction=bank_transaction,
                grant=grant,
                amount=bank_transaction.amount,
                allocation_type='source',
                created_by=request.user
            )

        return Response(
            {
                'success': True,
                'message': (
                    'Credit transaction assigned '
                    'to Grant successfully.'
                ),

                'transaction': {
                    'id': bank_transaction.id,
                    'date': bank_transaction.txn_date,
                    'type': bank_transaction.txn_type,
                    'amount': str(
                        bank_transaction.amount
                    )
                },

                'grant': {
                    'id': grant.id,
                    'name': grant.name,
                    'amount': str(grant.amount)
                },

                'allocation': {
                    'id': allocation.id,
                    'allocation_type': allocation.allocation_type,
                    'amount': str(allocation.amount)
                }
            },
            status=status.HTTP_201_CREATED
        )

    except Exception as e:

        logger.error(
            f'Error assigning transaction '
            f'{transaction_id} to grant: {e}',
            exc_info=True
        )

        return Response(
            {
                'success': False,
                'error': (
                    'Failed to assign credit transaction '
                    'to Grant.'
                )
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def get_transaction_receipt_info(request, transaction_id):

#     try:
#         bank_transaction = Transaction.objects.select_related(
#             'bank_account',
#             'bank_account__bank'
#         ).get(
#             id=transaction_id,
#             company=request.user.company
#         )
#         print(bank_transaction)

#     except Transaction.DoesNotExist:
#         return Response(
#             {
#                 'success': False,
#                 'error': 'Transaction not found'
#             },
#             status=status.HTTP_404_NOT_FOUND
#         )

#     # ---------------------------------------
#     # Get existing fund allocation
#     # ---------------------------------------

#     allocation = (
#         TransactionFundAllocation.objects
#         .select_related(
#             'fund',
#             'fund__grant'
#         )
#         .filter(
#             transaction=bank_transaction,
#             allocation_type='source'
#         )
#         .first()
#     )

#     grant_type = None
#     fund_name = None
#     grant_id = None
#     fund_id = None

#     if allocation and allocation.fund:

#         fund = allocation.fund

#         fund_id = fund.id
#         fund_name = fund.name

#         if fund.grant:
#             grant_id = fund.grant.id
#             grant_type = fund.grant.name


#     # -------------------------------------------------
#         # 3. Get receipts linked directly to this transaction
#         # -------------------------------------------------
#         receipts = ReceiptDocument.objects.filter(
#             matched_transaction=bank_transaction
#         ).order_by('-created_at')

#         # -------------------------------------------------
#         # 4. Calculate total of VALID / EXTRACTED receipts
#         # -------------------------------------------------
#         total_receipt_amount = (
#             receipts
#             .filter(
#                 extracted=True,
#                 amount__isnull=False
#             )
#             .aggregate(
#                 total=Sum('amount')
#             )['total']
#             or Decimal('0.00')
#         )

#         total_receipt_amount = Decimal(
#             str(total_receipt_amount)
#         )

#         transaction_amount = Decimal(
#             str(bank_transaction.amount or 0)
#         )

#         # -------------------------------------------------
#         # 5. Calculate remaining amount
#         # -------------------------------------------------
#         remaining_amount = (
#             transaction_amount - total_receipt_amount
#         )

#         # Don't allow negative remaining amount in the
#         # response in case old/incorrect data exists.
#         if remaining_amount < Decimal('0.00'):
#             remaining_amount = Decimal('0.00')

#         # -------------------------------------------------
#         # 6. Check whether receipt amount is complete
#         # -------------------------------------------------
#         is_receipt_amount_complete = (
#             total_receipt_amount == transaction_amount
#         )

#         # -------------------------------------------------
#         # 7. Receipt details
#         # -------------------------------------------------
#         receipt_data = []

#         for receipt in receipts:
#             receipt_data.append({
#                 'receipt_id': receipt.id,
#                 'file_name': receipt.file_name,
#                 'file_link': receipt.file_link,
#                 'amount': (
#                     str(receipt.amount)
#                     if receipt.amount is not None
#                     else None
#                 ),
#                 'receipt_no': receipt.receipt_no or None,
#                 'receipt_date': (
#                     str(receipt.receipt_date)
#                     if receipt.receipt_date
#                     else None
#                 ),
#                 'extracted': receipt.extracted,
#                 'error': receipt.extraction_error or None,
#                 'uploaded_at': (
#                     receipt.uploaded_at.isoformat()
#                     if receipt.uploaded_at
#                     else None
#                 ),
#             })

#     # ---------------------------------------
#     # Response
#     # ---------------------------------------

#     return Response(
#         {
#             'success': True,
#             'transaction': {
#                 'id': bank_transaction.id,

#                 'transaction_date': (
#                     bank_transaction.txn_date
#                     .strftime('%d-%m-%Y')
#                     if bank_transaction.txn_date
#                     else None
#                 ),

#                 'source_bank': (
#                     bank_transaction.bank_account.bank.bank_name
#                     if bank_transaction.bank_account
#                     else None
#                 ),

#                 'transaction_type': (
#                     bank_transaction.txn_type
#                 ),

#                 'reference_no': (
#                     bank_transaction.ref_no
#                 ),

#                 'amount': str(
#                     bank_transaction.amount
#                 ),

#                 "particular_name": (
#                     bank_transaction.particular.name
#                     if bank_transaction.particular
#                     else None
#                 ),

#                 # "particular_name": (
#                 #     transaction.particular.name
#                 #     if transaction.particular
#                 #     else None
#                 # ),

#                 'grant_id': grant_id,
#                 'grant_type': grant_type,

#                 'fund_id': fund_id,
#                 'fund_name': fund_name,

#                 'receipt_validation': {
#                         'amount': {
#                             'transaction_amount': str(
#                                 transaction_amount
#                             ),
#                             'total_receipt_amount': str(
#                                 total_receipt_amount
#                             ),
#                             'remaining_amount': str(
#                                 remaining_amount
#                             ),
#                             'is_complete': (
#                                 is_receipt_amount_complete
#                             ),
#                         }
#                     },
#                 'receipts': receipt_data,
#             }
#         },
#         status=status.HTTP_200_OK
#     )





# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def get_transaction_receipt_info(request, transaction_id):

#     # ---------------------------------------
#     # 1. Get transaction
#     # ---------------------------------------

#     try:
#         bank_transaction = (
#             Transaction.objects
#             .select_related(
#                 'bank_account',
#                 'bank_account__bank',
#                 'particular',
#             )
#             .get(
#                 id=transaction_id,
#                 company=request.user.company
#             )
#         )

#         print(bank_transaction)

#     except Transaction.DoesNotExist:
#         return Response(
#             {
#                 'success': False,
#                 'error': 'Transaction not found'
#             },
#             status=status.HTTP_404_NOT_FOUND
#         )

#     # ---------------------------------------
#     # 2. Get existing fund allocation
#     # ---------------------------------------

#     allocation = (
#         TransactionFundAllocation.objects
#         .select_related(
#             'fund',
#             'fund__grant'
#         )
#         .filter(
#             transaction=bank_transaction,
#             allocation_type='source'
#         )
#         .first()
#     )

#     # Default values
#     grant_type = None
#     fund_name = None
#     grant_id = None
#     fund_id = None

#     # ---------------------------------------
#     # Get fund / grant information
#     # ---------------------------------------

#     if allocation and allocation.fund:

#         fund = allocation.fund

#         fund_id = fund.id
#         fund_name = fund.name

#         if fund.grant:
#             grant_id = fund.grant.id
#             grant_type = fund.grant.name

#     # ---------------------------------------
#     # 3. Get receipts linked directly
#     #    to this transaction
#     # ---------------------------------------

#     receipts = (
#         ReceiptDocument.objects
#         .filter(
#             matched_transaction=bank_transaction
#         )
#         .order_by('-created_at')
#     )

#     # ---------------------------------------
#     # 4. Calculate total VALID / EXTRACTED
#     #    receipt amount
#     # ---------------------------------------

#     total_receipt_amount = (
#         receipts
#         .filter(
#             extracted=True,
#             amount__isnull=False
#         )
#         .aggregate(
#             total=Sum('amount')
#         )['total']
#         or Decimal('0.00')
#     )

#     total_receipt_amount = Decimal(
#         str(total_receipt_amount)
#     )

#     # ---------------------------------------
#     # 5. Get transaction amount
#     # ---------------------------------------

#     transaction_amount = Decimal(
#         str(bank_transaction.amount or 0)
#     )

#     # ---------------------------------------
#     # 6. Calculate remaining amount
#     # ---------------------------------------

#     remaining_amount = (
#         transaction_amount - total_receipt_amount
#     )

#     # Don't allow negative remaining amount
#     if remaining_amount < Decimal('0.00'):
#         remaining_amount = Decimal('0.00')

#     # ---------------------------------------
#     # 7. Check whether receipt amount
#     #    is complete
#     # ---------------------------------------

#     is_receipt_amount_complete = (
#         total_receipt_amount == transaction_amount
#     )

#     # ---------------------------------------
#     # 8. Receipt details
#     # ---------------------------------------

#     receipt_data = []

#     for receipt in receipts:

#         receipt_data.append({
#             'receipt_id': receipt.id,

#             'file_name': receipt.file_name,

#             'file_link': receipt.file_link,

#             'amount': (
#                 str(receipt.amount)
#                 if receipt.amount is not None
#                 else None
#             ),

#             'receipt_no': (
#                 receipt.receipt_no
#                 if receipt.receipt_no
#                 else None
#             ),

#             'receipt_date': (
#                 str(receipt.receipt_date)
#                 if receipt.receipt_date
#                 else None
#             ),

#             'extracted': receipt.extracted,

#             'error': (
#                 receipt.extraction_error
#                 if receipt.extraction_error
#                 else None
#             ),

#             'uploaded_at': (
#                 receipt.uploaded_at.isoformat()
#                 if receipt.uploaded_at
#                 else None
#             ),
#         })

#     # ---------------------------------------
#     # 9. Response
#     # ---------------------------------------

#     return Response(
#         {
#             'success': True,

#             'transaction': {

#                 'id': bank_transaction.id,

#                 'transaction_date': (
#                     bank_transaction.txn_date.strftime('%d-%m-%Y')
#                     if bank_transaction.txn_date
#                     else None
#                 ),

#                 'source_bank': (
#                     bank_transaction.bank_account.bank.bank_name
#                     if bank_transaction.bank_account
#                     and bank_transaction.bank_account.bank
#                     else None
#                 ),

#                 'transaction_type': (
#                     bank_transaction.txn_type
#                 ),

#                 'reference_no': (
#                     bank_transaction.ref_no
#                 ),

#                 'amount': str(
#                     bank_transaction.amount
#                     if bank_transaction.amount is not None
#                     else Decimal('0.00')
#                 ),

#                 'particular_name': (
#                     bank_transaction.particular.name
#                     if bank_transaction.particular
#                     else None
#                 ),

#                 # ---------------------------------------
#                 # Fund / Grant
#                 # ---------------------------------------

#                 'grant_id': grant_id,

#                 'grant_type': grant_type,

#                 'fund_id': fund_id,

#                 'fund_name': fund_name,

#                 # ---------------------------------------
#                 # Receipt validation
#                 # ---------------------------------------

#                 'receipt_validation': {

#                     'amount': {

#                         'transaction_amount': str(
#                             transaction_amount
#                         ),

#                         'total_receipt_amount': str(
#                             total_receipt_amount
#                         ),

#                         'remaining_amount': str(
#                             remaining_amount
#                         ),

#                         'is_complete': (
#                             is_receipt_amount_complete
#                         ),
#                     }
#                 },

#                 # ---------------------------------------
#                 # Receipts
#                 # ---------------------------------------

#                 'receipts': receipt_data,
#             }
#         },
#         status=status.HTTP_200_OK
#     )



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_transaction_receipt_info(request, transaction_id):

    # ---------------------------------------
    # 1. Get transaction
    # ---------------------------------------

    try:
        bank_transaction = (
            Transaction.objects
            .select_related(
                'bank_account',
                'bank_account__bank',
                'particular',
            )
            .get(
                id=transaction_id,
                company=request.user.company
            )
        )

    except Transaction.DoesNotExist:
        return Response(
            {
                'success': False,
                'error': 'Transaction not found'
            },
            status=status.HTTP_404_NOT_FOUND
        )


    transaction_type = bank_transaction.txn_type.lower()

    if transaction_type == 'credit':

        allocation_type = 'source'

    elif transaction_type == 'debit':

        allocation_type = 'destination'

    else:

        return Response(
            {
                'success': False,
                'error': 'Invalid transaction type.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    # ---------------------------------------
    # 2. Get Grant allocation
    # ---------------------------------------

    allocation = (
        TransactionGrantAllocation.objects
        .select_related(
            'grant',
            'grant__agency'
        )
        .filter(
            transaction=bank_transaction,
            allocation_type=allocation_type
        )
        .first()
    )

    # Default values
    grant_id = None
    grant_name = None
    agency_id = None
    agency_name = None
    grant_amount = None
    allocation_id = None

    # ---------------------------------------
    # 3. Get Grant / Agency information
    # ---------------------------------------

    if allocation and allocation.grant:

        grant = allocation.grant

        allocation_id = allocation.id
        grant_id = grant.id
        grant_name = grant.name
        grant_amount = grant.amount

        if grant.agency:
            agency_id = grant.agency.id
            agency_name = grant.agency.name

    # ---------------------------------------
    # 4. Get receipts linked directly
    #    to this transaction
    # ---------------------------------------

    receipts = (
        ReceiptDocument.objects
        .filter(
            matched_transaction=bank_transaction,
            extraction_error=''
        )
        .order_by('-created_at')
    )

    # ---------------------------------------
    # 5. Calculate total VALID / EXTRACTED
    #    receipt amount
    # ---------------------------------------

    total_receipt_amount = (
        receipts
        .filter(
            extracted=True,
            amount__isnull=False
        )
        .aggregate(
            total=Sum('amount')
        )['total']
        or Decimal('0.00')
    )

    total_receipt_amount = Decimal(
        str(total_receipt_amount)
    )

    # ---------------------------------------
    # 6. Get transaction amount
    # ---------------------------------------

    transaction_amount = Decimal(
        str(bank_transaction.amount or 0)
    )

    # ---------------------------------------
    # 7. Calculate remaining receipt amount
    # ---------------------------------------

    remaining_amount = (
        transaction_amount - total_receipt_amount
    )

    if remaining_amount < Decimal('0.00'):
        remaining_amount = Decimal('0.00')

    # ---------------------------------------
    # 8. Check receipt amount completion
    # ---------------------------------------

    is_receipt_amount_complete = (
        total_receipt_amount == transaction_amount
    )

    # ---------------------------------------
    # 9. Receipt details
    # ---------------------------------------

    receipt_data = []

    for receipt in receipts:

        receipt_data.append({
            'receipt_id': receipt.id,

            'file_name': receipt.file_name,

            'file_link': receipt.file_link,

            'amount': (
                str(receipt.amount)
                if receipt.amount is not None
                else None
            ),

            'receipt_no': (
                receipt.receipt_no
                if receipt.receipt_no
                else None
            ),

            'receipt_date': (
                str(receipt.receipt_date)
                if receipt.receipt_date
                else None
            ),

            'extracted': receipt.extracted,

            'error': (
                receipt.extraction_error
                if receipt.extraction_error
                else None
            ),

            'uploaded_at': (
                receipt.uploaded_at.isoformat()
                if receipt.uploaded_at
                else None
            ),
        })

    # ---------------------------------------
    # 10. Response
    # ---------------------------------------

    return Response(
        {
            'success': True,

            'transaction': {

                'id': bank_transaction.id,

                'transaction_date': (
                    bank_transaction.txn_date.strftime('%d-%m-%Y')
                    if bank_transaction.txn_date
                    else None
                ),

                'source_bank': (
                    bank_transaction.bank_account.bank.bank_name
                    if bank_transaction.bank_account
                    and bank_transaction.bank_account.bank
                    else None
                ),

                'transaction_type': (
                    bank_transaction.txn_type
                ),

                'reference_no': (
                    bank_transaction.ref_no
                ),

                'amount': str(
                    bank_transaction.amount
                    if bank_transaction.amount is not None
                    else Decimal('0.00')
                ),

                'particular_name': (
                    bank_transaction.particular.name
                    if bank_transaction.particular
                    else None
                ),

                # ---------------------------------------
                # Agency
                # ---------------------------------------

                'agency_id': agency_id,

                'agency_name': agency_name,

                # ---------------------------------------
                # Grant
                # ---------------------------------------

                'grant_id': grant_id,

                'grant_name': grant_name,

                'grant_amount': (
                    str(grant_amount)
                    if grant_amount is not None
                    else None
                ),

                'allocation_id': allocation_id,

                # ---------------------------------------
                # Receipt validation
                # ---------------------------------------

                'receipt_validation': {

                    'amount': {

                        'transaction_amount': str(
                            transaction_amount
                        ),

                        'total_receipt_amount': str(
                            total_receipt_amount
                        ),

                        'remaining_amount': str(
                            remaining_amount
                        ),

                        'is_complete': (
                            is_receipt_amount_complete
                        ),
                    }
                },

                # ---------------------------------------
                # Receipts
                # ---------------------------------------

                'receipts': receipt_data,
            }
        },
        status=status.HTTP_200_OK
    )



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_company_grants(request):
    company = request.user.company

    if not company:
        return Response(
            {
                'success': False,
                'error': 'User is not associated with a company.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    grants = Grant.objects.filter(
        company=company,
        is_active=True
    ).order_by('name')

    serializer = GrantListSerializer(
        grants,
        many=True
    )

    return Response(
        {
            'success': True,
            'count': grants.count(),
            'grants': serializer.data
        },
        status=status.HTTP_200_OK
    )





def get_fund_available_amount(fund, bank_account):
    """
    Calculate the current amount of a Fund available
    in a particular Bank Account.
    """

    allocations = (
        TransactionFundAllocation.objects
        .filter(
            fund=fund,
            transaction__bank_account=bank_account
        )
        .select_related('transaction')
    )

    available_amount = Decimal('0.00')

    for allocation in allocations:

        txn = allocation.transaction

        if txn.txn_type.lower() == 'credit':
            available_amount += allocation.amount

        elif txn.txn_type.lower() == 'debit':
            available_amount -= allocation.amount

    return available_amount


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def get_available_funds_for_transaction(
#     request,
#     transaction_id
# ):

#     company = request.user.company

#     # =====================================================
#     # 1. Validate company
#     # =====================================================

#     if not company:
#         return Response(
#             {
#                 'success': False,
#                 'error': 'User is not associated with a company.'
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     # =====================================================
#     # 2. Get transaction
#     # =====================================================

#     try:

#         bank_transaction = (
#             Transaction.objects
#             .select_related('bank_account')
#             .get(
#                 id=transaction_id,
#                 company=company
#             )
#         )

#     except Transaction.DoesNotExist:

#         return Response(
#             {
#                 'success': False,
#                 'error': 'Transaction not found.'
#             },
#             status=status.HTTP_404_NOT_FOUND
#         )

#     # =====================================================
#     # 3. Only Debit transactions
#     # =====================================================

#     if bank_transaction.txn_type.lower() != 'debit':

#         return Response(
#             {
#                 'success': False,
#                 'error': (
#                     'Available Funds can only be requested '
#                     'for debit transactions.'
#                 )
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     bank_account = bank_transaction.bank_account

#     # =====================================================
#     # 4. Optional Grant filter
#     # =====================================================

#     grant_id = request.query_params.get('grant_id')

#     # =====================================================
#     # 5. No Grant selected
#     #
#     #    Return ONLY Grants having a Fund available
#     #    in THIS transaction's Bank Account.
#     # =====================================================

#     if not grant_id:

#         funds = (
#             Fund.objects
#             .select_related('grant')
#             .filter(
#                 grant__company=company,
#                 grant__is_active=True
#             )
#             .order_by(
#                 'grant__name',
#                 'name'
#             )
#         )

#         available_grant_ids = set()

#         for fund in funds:

#             available_amount = get_fund_available_amount(
#                 fund=fund,
#                 bank_account=bank_account
#             )

#             if available_amount > Decimal('0.00'):

#                 available_grant_ids.add(
#                     fund.grant_id
#                 )

#         grants = (
#             Grant.objects
#             .filter(
#                 id__in=available_grant_ids,
#                 company=company,
#                 is_active=True
#             )
#             .order_by('name')
#         )

#         return Response(
#             {
#                 'success': True,

#                 'transaction': {
#                     'id': bank_transaction.id,
#                     'bank_account_id': bank_account.id,
#                     'amount': str(
#                         bank_transaction.amount
#                     )
#                 },

#                 'grants': [
#                     {
#                         'id': grant.id,
#                         'name': grant.name
#                     }
#                     for grant in grants
#                 ]
#             },
#             status=status.HTTP_200_OK
#         )

#     # =====================================================
#     # 6. Grant selected
#     # =====================================================

#     try:

#         grant = Grant.objects.get(
#             id=grant_id,
#             company=company,
#             is_active=True
#         )

#     except Grant.DoesNotExist:

#         return Response(
#             {
#                 'success': False,
#                 'error': 'Grant not found or inactive.'
#             },
#             status=status.HTTP_404_NOT_FOUND
#         )

#     # =====================================================
#     # 7. Get ONLY this Grant's Funds
#     # =====================================================

#     funds = (
#         Fund.objects
#         .filter(
#             grant=grant
#         )
#         .order_by('name')
#     )

#     available_funds = []

#     # =====================================================
#     # 8. Check each Fund against THIS Bank Account
#     # =====================================================

#     for fund in funds:

#         available_amount = get_fund_available_amount(
#             fund=fund,
#             bank_account=bank_account
#         )

#         if available_amount <= Decimal('0.00'):
#             continue

#         available_funds.append(
#             {
#                 'id': fund.id,
#                 'name': fund.name,
#                 'available_amount': str(
#                     available_amount
#                 )
#             }
#         )

#     # =====================================================
#     # 9. Response
#     # =====================================================

#     return Response(
#         {
#             'success': True,

#             'transaction': {
#                 'id': bank_transaction.id,
#                 'bank_account_id': bank_account.id,
#                 'amount': str(
#                     bank_transaction.amount
#                 )
#             },

#             'grant': {
#                 'id': grant.id,
#                 'name': grant.name
#             },

#             'funds': available_funds
#         },
#         status=status.HTTP_200_OK
#     )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_available_grants_for_transaction(
    request,
    transaction_id
):

    company = request.user.company

    # =====================================================
    # 1. Validate company
    # =====================================================

    if not company:
        return Response(
            {
                'success': False,
                'error': 'User is not associated with a company.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # =====================================================
    # 2. Get transaction
    # =====================================================

    try:

        bank_transaction = (
            Transaction.objects
            .select_related(
                'bank_account',
                'bank_account__bank'
            )
            .get(
                id=transaction_id,
                company=company
            )
        )

    except Transaction.DoesNotExist:

        return Response(
            {
                'success': False,
                'error': 'Transaction not found.'
            },
            status=status.HTTP_404_NOT_FOUND
        )

    # =====================================================
    # 3. Only Debit transactions
    # =====================================================

    if bank_transaction.txn_type.lower() != 'debit':

        return Response(
            {
                'success': False,
                'error': (
                    'Available Grants can only be requested '
                    'for debit transactions.'
                )
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    bank_account = bank_transaction.bank_account

    # =====================================================
    # 4. Optional Grant filter
    # =====================================================

    grant_id = request.query_params.get('grant_id')

    # =====================================================
    # 5. No Grant selected
    #
    # Return ONLY Grants having available amount
    # in THIS transaction's Bank Account.
    # =====================================================

    if not grant_id:

        grants = (
            Grant.objects
            .select_related('agency')
            .filter(
                company=company,
                is_active=True
            )
            .order_by(
                'agency__name',
                'name'
            )
        )

        available_grants = []

        for grant in grants:

            available_amount = get_grant_available_amount(
                grant=grant,
                bank_account=bank_account
            )

            if available_amount <= Decimal('0.00'):
                continue

            available_grants.append(
                {
                    'id': grant.id,
                    'name': grant.name,

                    'agency_id': (
                        grant.agency.id
                        if grant.agency
                        else None
                    ),

                    'agency_name': (
                        grant.agency.name
                        if grant.agency
                        else None
                    ),

                    'available_amount': str(
                        available_amount
                    )
                }
            )

        return Response(
            {
                'success': True,

                'transaction': {
                    'id': bank_transaction.id,

                    'bank_account_id': (
                        bank_account.id
                    ),

                    'amount': str(
                        bank_transaction.amount
                    )
                },

                'grants': available_grants
            },
            status=status.HTTP_200_OK
        )

    # =====================================================
    # 6. Grant selected
    # =====================================================

    try:

        grant = (
            Grant.objects
            .select_related('agency')
            .get(
                id=grant_id,
                company=company,
                is_active=True
            )
        )

    except Grant.DoesNotExist:

        return Response(
            {
                'success': False,
                'error': 'Grant not found or inactive.'
            },
            status=status.HTTP_404_NOT_FOUND
        )

    # =====================================================
    # 7. Calculate available amount for this Grant
    #    in THIS Bank Account
    # =====================================================

    available_amount = get_grant_available_amount(
        grant=grant,
        bank_account=bank_account
    )

    # =====================================================
    # 8. No available amount
    # =====================================================

    if available_amount <= Decimal('0.00'):

        return Response(
            {
                'success': True,

                'transaction': {
                    'id': bank_transaction.id,
                    'bank_account_id': bank_account.id,
                    'amount': str(
                        bank_transaction.amount
                    )
                },

                'grant': {
                    'id': grant.id,
                    'name': grant.name,

                    'agency_id': (
                        grant.agency.id
                        if grant.agency
                        else None
                    ),

                    'agency_name': (
                        grant.agency.name
                        if grant.agency
                        else None
                    ),

                    'available_amount': '0.00'
                }
            },
            status=status.HTTP_200_OK
        )

    # =====================================================
    # 9. Response
    # =====================================================

    return Response(
        {
            'success': True,

            'transaction': {
                'id': bank_transaction.id,

                'bank_account_id': (
                    bank_account.id
                ),

                'amount': str(
                    bank_transaction.amount
                )
            },

            'grant': {
                'id': grant.id,
                'name': grant.name,

                'agency_id': (
                    grant.agency.id
                    if grant.agency
                    else None
                ),

                'agency_name': (
                    grant.agency.name
                    if grant.agency
                    else None
                ),

                'available_amount': str(
                    available_amount
                )
            }
        },
        status=status.HTTP_200_OK
    )




# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def assign_funds_to_debit(request, transaction_id):

#     company = request.user.company

#     # =========================================================
#     # 1. Validate company
#     # =========================================================

#     if not company:
#         return Response(
#             {
#                 'success': False,
#                 'error': 'User is not associated with a company.'
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     # =========================================================
#     # 2. Get transaction
#     # =========================================================

#     try:
#         bank_transaction = (
#             Transaction.objects.select_related(
#                 'bank_account'
#             )
#             .get(
#                 id=transaction_id,
#                 company=company
#             )
#         )

#     except Transaction.DoesNotExist:

#         return Response(
#             {
#                 'success': False,
#                 'error': 'Transaction not found.'
#             },
#             status=status.HTTP_404_NOT_FOUND
#         )

#     # =========================================================
#     # 3. Only Debit transactions
#     # =========================================================

#     if bank_transaction.txn_type.lower() != 'debit':

#         return Response(
#             {
#                 'success': False,
#                 'error': (
#                     'Funds can only be assigned '
#                     'to debit transactions.'
#                 )
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     transaction_amount = bank_transaction.amount

#     # =========================================================
#     # 4. Get allocations from request
#     # =========================================================

#     allocations = request.data.get('allocations')

#     if not allocations:

#         return Response(
#             {
#                 'success': False,
#                 'error': 'At least one fund must be selected.'
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     if not isinstance(allocations, list):

#         return Response(
#             {
#                 'success': False,
#                 'error': 'allocations must be a list.'
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     # =========================================================
#     # 5. Validate number of selected Funds
#     # =========================================================

#     if len(allocations) == 1:
#         fund_id = allocations[0].get('fund_id')
#         if not fund_id:

#             return Response(
#                 {
#                     'success': False,
#                     'error': 'fund_id is required.'
#                 },
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#         # -----------------------------------------------------
#         # One Fund selected:
#         # automatically use complete transaction amount
#         # -----------------------------------------------------

#         allocations[0]['amount'] = transaction_amount

#     else:

#         # -----------------------------------------------------
#         # Multiple Funds:
#         # amount is mandatory for every Fund
#         # -----------------------------------------------------

#         for allocation in allocations:

#             if not allocation.get('fund_id'):

#                 return Response(
#                     {
#                         'success': False,
#                         'error': 'fund_id is required for every allocation.'
#                     },
#                     status=status.HTTP_400_BAD_REQUEST
#                 )

#             if allocation.get('amount') is None:

#                 return Response(
#                     {
#                         'success': False,
#                         'error': (
#                             'Amount is required when '
#                             'multiple funds are selected.'
#                         )
#                     },
#                     status=status.HTTP_400_BAD_REQUEST
#                 )

#     # =========================================================
#     # 6. Validate amounts and prepare allocations
#     # =========================================================

#     prepared_allocations = []

#     total_allocation = Decimal('0.00')

#     seen_funds = set()

#     for item in allocations:

#         fund_id = item.get('fund_id')

#         # -----------------------------------------------------
#         # Prevent same Fund twice
#         # -----------------------------------------------------

#         if fund_id in seen_funds:

#             return Response(
#                 {
#                     'success': False,
#                     'error': (
#                         f'Fund {fund_id} is selected more than once.'
#                     )
#                 },
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#         seen_funds.add(fund_id)

#         # -----------------------------------------------------
#         # Convert amount
#         # -----------------------------------------------------

#         try:

#             amount = Decimal(
#                 str(item.get('amount'))
#             )

#         except (InvalidOperation, TypeError, ValueError):

#             return Response(
#                 {
#                     'success': False,
#                     'error': (
#                         f'Invalid amount for fund {fund_id}.'
#                     )
#                 },
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#         if amount <= 0:

#             return Response(
#                 {
#                     'success': False,
#                     'error': (
#                         f'Amount for fund {fund_id} '
#                         'must be greater than zero.'
#                     )
#                 },
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#         total_allocation += amount

#         prepared_allocations.append(
#             {
#                 'fund_id': fund_id,
#                 'amount': amount
#             }
#         )

#     # =========================================================
#     # 7. Total allocation must equal transaction amount
#     # =========================================================

#     if total_allocation != transaction_amount:

#         return Response(
#             {
#                 'success': False,
#                 'error': (
#                     'Total fund allocation must equal '
#                     'the transaction amount.'
#                 ),
#                 'transaction_amount': str(transaction_amount),
#                 'allocated_amount': str(total_allocation)
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     # =========================================================
#     # 8. Check whether this is an Internal Transfer
#     # =========================================================

#     destination_transaction = None

#     if bank_transaction.matched_txn_id:

#         try:

#             destination_transaction = (
#                 Transaction.objects.select_related('bank_account')
#                 .get(
#                     id=bank_transaction.matched_txn_id,
#                     company=company
#                 )
#             )

#         except Transaction.DoesNotExist:

#             return Response(
#                 {
#                     'success': False,
#                     'error': (
#                         'Matched transaction was not found.'
#                     )
#                 },
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#         # -----------------------------------------------------
#         # Matched transaction must be Credit
#         # -----------------------------------------------------

#         if destination_transaction.txn_type.lower() != 'credit':

#             return Response(
#                 {
#                     'success': False,
#                     'error': (
#                         'The matched transaction is not a credit '
#                         'transaction.'
#                     )
#                 },
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#         # -----------------------------------------------------
#         # Amount must match
#         # -----------------------------------------------------

#         if destination_transaction.amount != transaction_amount:

#             return Response(
#                 {
#                     'success': False,
#                     'error': (
#                         'Debit and matched credit amounts '
#                         'do not match.'
#                     )
#                 },
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#         # -----------------------------------------------------
#         # Must be different bank accounts
#         # -----------------------------------------------------

#         if (
#             bank_transaction.bank_account_id
#             == destination_transaction.bank_account_id
#         ):

#             return Response(
#                 {
#                     'success': False,
#                     'error': (
#                         'Internal transfer must involve '
#                         'different bank accounts.'
#                     )
#                 },
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#     # =========================================================
#     # 9. Validate Funds
#     # =========================================================

#     funds = {}

#     for item in prepared_allocations:

#         try:

#             fund = (
#                 Fund.objects.select_related('grant')
#                 .get(
#                     id=item['fund_id'],
#                     grant__company=company
#                 )
#             )

#         except Fund.DoesNotExist:

#             return Response(
#                 {
#                     'success': False,
#                     'error': (
#                         f"Fund {item['fund_id']} "
#                         "not found."
#                     )
#                 },
#                 status=status.HTTP_404_NOT_FOUND
#             )

#         funds[item['fund_id']] = fund

#     # =========================================================
#     # 10. Calculate available amount per Fund in this Bank
#     # =========================================================

#     for item in prepared_allocations:

#         fund = funds[item['fund_id']]
#         requested_amount = item['amount']

#         allocations_for_bank = (
#             TransactionFundAllocation.objects
#             .filter(
#                 fund=fund,
#                 transaction__bank_account=(
#                     bank_transaction.bank_account
#                 )
#             )
#             .select_related('transaction')
#         )

#         available_amount = Decimal('0.00')

#         for allocation in allocations_for_bank:

#             txn = allocation.transaction

#             if txn.txn_type.lower() == 'credit':

#                 available_amount += allocation.amount

#             elif txn.txn_type.lower() == 'debit':

#                 available_amount -= allocation.amount

#         if requested_amount > available_amount:

#             return Response(
#                 {
#                     'success': False,
#                     'error': (
#                         f'Insufficient amount available '
#                         f'for Fund "{fund.name}".'
#                     ),
#                     'fund_id': fund.id,
#                     'fund_name': fund.name,
#                     'available_amount': str(
#                         available_amount
#                     ),
#                     'requested_amount': str(
#                         requested_amount
#                     )
#                 },
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#     # =========================================================
#     # 11. Prevent duplicate allocation on source transaction
#     # =========================================================

#     if TransactionFundAllocation.objects.filter(
#         transaction=bank_transaction
#     ).exists():

#         return Response(
#             {
#                 'success': False,
#                 'error': (
#                     'Funds have already been assigned '
#                     'to this debit transaction.'
#                 )
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     # =========================================================
#     # 12. Internal Transfer:
#     # destination must not already have allocations
#     # =========================================================

#     if destination_transaction:

#         if TransactionFundAllocation.objects.filter(
#             transaction=destination_transaction
#         ).exists():

#             return Response(
#                 {
#                     'success': False,
#                     'error': (
#                         'The matched credit transaction '
#                         'already has Fund allocations.'
#                     )
#                 },
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#     # =========================================================
#     # 13. Save everything atomically
#     # =========================================================

#     try:

#         with db_transaction.atomic():

#             internal_transfer = None

#             # -------------------------------------------------
#             # Create InternalTransfer only when matched_txn
#             # exists
#             # -------------------------------------------------

#             if destination_transaction:

#                 internal_transfer = (
#                     InternalTransfer.objects.create(
#                         company=company,
#                         transfer_date=(
#                             bank_transaction.txn_date
#                         ),
#                         amount=transaction_amount,
#                         created_by=request.user
#                     )
#                 )

#                 InternalTransferTransaction.objects.create(
#                     internal_transfer=internal_transfer,
#                     transaction=bank_transaction,
#                     role='source'
#                 )

#                 InternalTransferTransaction.objects.create(
#                     internal_transfer=internal_transfer,
#                     transaction=destination_transaction,
#                     role='destination'
#                 )

#             created_source_allocations = []
#             created_destination_allocations = []

#             # -------------------------------------------------
#             # Create source allocations
#             # -------------------------------------------------

#             for item in prepared_allocations:

#                 fund = funds[item['fund_id']]
#                 amount = item['amount']

#                 source_allocation = (
#                     TransactionFundAllocation.objects.create(
#                         transaction=bank_transaction,
#                         fund=fund,
#                         amount=amount,
#                         allocation_type='source',
#                         created_by=request.user
#                     )
#                 )

#                 created_source_allocations.append(
#                     source_allocation
#                 )

#                 # ---------------------------------------------
#                 # Internal transfer:
#                 # automatically assign same Fund to
#                 # destination credit
#                 # ---------------------------------------------

#                 if destination_transaction:

#                     destination_allocation = (
#                         TransactionFundAllocation.objects.create(
#                             transaction=destination_transaction,
#                             fund=fund,
#                             amount=amount,
#                             allocation_type='destination',
#                             created_by=request.user
#                         )
#                     )

#                     created_destination_allocations.append(
#                         destination_allocation
#                     )

#         # =====================================================
#         # Response
#         # =====================================================

#         response_data = {
#             'success': True,
#             'message': (
#                 'Funds assigned successfully.'
#             ),
#             'transaction': {
#                 'id': bank_transaction.id,
#                 'amount': str(transaction_amount),
#                 'transaction_type': bank_transaction.txn_type
#             },
#             'allocations': [
#                 {
#                     'fund_id': allocation.fund_id,
#                     'fund_name': allocation.fund.name,
#                     'amount': str(allocation.amount)
#                 }
#                 for allocation
#                 in created_source_allocations
#             ]
#         }

#         # -----------------------------------------------------
#         # Internal transfer response
#         # -----------------------------------------------------

#         if destination_transaction:

#             response_data['internal_transfer'] = {
#                 'id': internal_transfer.id,
#                 'source_transaction_id': (
#                     bank_transaction.id
#                 ),
#                 'destination_transaction_id': (
#                     destination_transaction.id
#                 ),
#                 'amount': str(transaction_amount)
#             }

#             response_data['destination'] = {
#                 'transaction_id': (
#                     destination_transaction.id
#                 ),
#                 'allocations': [
#                     {
#                         'fund_id': allocation.fund_id,
#                         'fund_name': allocation.fund.name,
#                         'amount': str(allocation.amount)
#                     }
#                     for allocation
#                     in created_destination_allocations
#                 ]
#             }

#         return Response(
#             response_data,
#             status=status.HTTP_201_CREATED
#         )

#     except Exception as e:

#         return Response(
#             {
#                 'success': False,
#                 'error': str(e)
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def assign_grants_to_debit(request, transaction_id):

    company = request.user.company

    # =========================================================
    # 1. Validate company
    # =========================================================

    if not company:
        return Response(
            {
                'success': False,
                'error': 'User is not associated with a company.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # =========================================================
    # 2. Get transaction
    # =========================================================

    try:

        bank_transaction = (
            Transaction.objects
            .select_related('bank_account')
            .get(
                id=transaction_id,
                company=company
            )
        )

    except Transaction.DoesNotExist:

        return Response(
            {
                'success': False,
                'error': 'Transaction not found.'
            },
            status=status.HTTP_404_NOT_FOUND
        )

    # =========================================================
    # 3. Only Debit transactions
    # =========================================================

    if bank_transaction.txn_type.lower() != 'debit':

        return Response(
            {
                'success': False,
                'error': (
                    'Grants can only be assigned '
                    'to debit transactions.'
                )
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    transaction_amount = bank_transaction.amount

    if (
        transaction_amount is None
        or transaction_amount <= Decimal('0.00')
    ):
        return Response(
            {
                'success': False,
                'error': (
                    'Debit transaction amount must '
                    'be greater than zero.'
                )
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # =========================================================
    # 4. Get allocations from request
    # =========================================================

    allocations = request.data.get('allocations')

    if not allocations:

        return Response(
            {
                'success': False,
                'error': 'At least one Grant must be selected.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    if not isinstance(allocations, list):

        return Response(
            {
                'success': False,
                'error': 'allocations must be a list.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # =========================================================
    # 5. Validate number of selected Grants
    # =========================================================

    if len(allocations) == 1:

        grant_id = allocations[0].get('grant_id')

        if not grant_id:

            return Response(
                {
                    'success': False,
                    'error': 'grant_id is required.'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # -----------------------------------------------------
        # One Grant:
        # automatically use complete transaction amount
        # -----------------------------------------------------

        allocations[0]['amount'] = transaction_amount

    else:

        # -----------------------------------------------------
        # Multiple Grants:
        # amount is mandatory for every Grant
        # -----------------------------------------------------

        for allocation in allocations:

            if not allocation.get('grant_id'):

                return Response(
                    {
                        'success': False,
                        'error': (
                            'grant_id is required for '
                            'every allocation.'
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            if allocation.get('amount') is None:

                return Response(
                    {
                        'success': False,
                        'error': (
                            'Amount is required when '
                            'multiple Grants are selected.'
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

    # =========================================================
    # 6. Validate amounts and prepare allocations
    # =========================================================

    prepared_allocations = []

    total_allocation = Decimal('0.00')

    seen_grants = set()

    for item in allocations:

        grant_id = item.get('grant_id')

        # -----------------------------------------------------
        # Prevent same Grant twice
        # -----------------------------------------------------

        if grant_id in seen_grants:

            return Response(
                {
                    'success': False,
                    'error': (
                        f'Grant {grant_id} is selected '
                        'more than once.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        seen_grants.add(grant_id)

        # -----------------------------------------------------
        # Convert amount
        # -----------------------------------------------------

        try:

            amount = Decimal(
                str(item.get('amount'))
            )

        except (
            InvalidOperation,
            TypeError,
            ValueError
        ):

            return Response(
                {
                    'success': False,
                    'error': (
                        f'Invalid amount for Grant '
                        f'{grant_id}.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        if amount <= Decimal('0.00'):

            return Response(
                {
                    'success': False,
                    'error': (
                        f'Amount for Grant {grant_id} '
                        'must be greater than zero.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        total_allocation += amount

        prepared_allocations.append(
            {
                'grant_id': grant_id,
                'amount': amount
            }
        )

    # =========================================================
    # 7. Total allocation must equal transaction amount
    # =========================================================

    if total_allocation != transaction_amount:

        return Response(
            {
                'success': False,
                'error': (
                    'Total Grant allocation must equal '
                    'the transaction amount.'
                ),
                'transaction_amount': str(
                    transaction_amount
                ),
                'allocated_amount': str(
                    total_allocation
                )
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # =========================================================
    # 8. Check whether this is an Internal Transfer
    # =========================================================

    destination_transaction = None

    if bank_transaction.matched_txn_id:

        try:

            destination_transaction = (
                Transaction.objects
                .select_related('bank_account')
                .get(
                    id=bank_transaction.matched_txn_id,
                    company=company
                )
            )

        except Transaction.DoesNotExist:

            return Response(
                {
                    'success': False,
                    'error': (
                        'Matched transaction was not found.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # -----------------------------------------------------
        # Matched transaction must be Credit
        # -----------------------------------------------------

        if destination_transaction.txn_type.lower() != 'credit':

            return Response(
                {
                    'success': False,
                    'error': (
                        'The matched transaction is not '
                        'a credit transaction.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # -----------------------------------------------------
        # Amount must match
        # -----------------------------------------------------

        if destination_transaction.amount != transaction_amount:

            return Response(
                {
                    'success': False,
                    'error': (
                        'Debit and matched credit amounts '
                        'do not match.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # -----------------------------------------------------
        # Must be different bank accounts
        # -----------------------------------------------------

        if (
            bank_transaction.bank_account_id
            == destination_transaction.bank_account_id
        ):

            return Response(
                {
                    'success': False,
                    'error': (
                        'Internal transfer must involve '
                        'different bank accounts.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

    # =========================================================
    # 9. Validate Grants
    # =========================================================

    grants = {}

    for item in prepared_allocations:

        try:

            grant = (
                Grant.objects
                .select_related('agency')
                .get(
                    id=item['grant_id'],
                    company=company,
                    is_active=True
                )
            )

        except Grant.DoesNotExist:

            return Response(
                {
                    'success': False,
                    'error': (
                        f"Grant {item['grant_id']} "
                        "not found or inactive."
                    )
                },
                status=status.HTTP_404_NOT_FOUND
            )

        grants[item['grant_id']] = grant

    # =========================================================
    # 10. Calculate available amount per Grant
    #     in THIS Bank Account
    # =========================================================

    for item in prepared_allocations:

        grant = grants[item['grant_id']]
        requested_amount = item['amount']

        available_amount = get_grant_available_amount(
            grant=grant,
            bank_account=bank_transaction.bank_account
        )

        if requested_amount > available_amount:

            return Response(
                {
                    'success': False,
                    'error': (
                        f'Insufficient amount available '
                        f'for Grant "{grant.name}".'
                    ),
                    'grant_id': grant.id,
                    'grant_name': grant.name,
                    'available_amount': str(
                        available_amount
                    ),
                    'requested_amount': str(
                        requested_amount
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

    # =========================================================
    # 11. Prevent duplicate allocation on debit transaction
    # =========================================================

    if TransactionGrantAllocation.objects.filter(
        transaction=bank_transaction
    ).exists():

        return Response(
            {
                'success': False,
                'error': (
                    'Grants have already been assigned '
                    'to this debit transaction.'
                )
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # =========================================================
    # 12. Internal Transfer:
    #     destination must not already have allocations
    # =========================================================

    if destination_transaction:

        if TransactionGrantAllocation.objects.filter(
            transaction=destination_transaction
        ).exists():

            return Response(
                {
                    'success': False,
                    'error': (
                        'The matched credit transaction '
                        'already has Grant allocations.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

    # =========================================================
    # 13. Save everything atomically
    # =========================================================

    try:

        with db_transaction.atomic():

            internal_transfer = None

            # -------------------------------------------------
            # Create InternalTransfer when matched transaction
            # exists
            # -------------------------------------------------

            if destination_transaction:

                internal_transfer = (
                    InternalTransfer.objects.create(
                        company=company,
                        transfer_date=(
                            bank_transaction.txn_date
                        ),
                        amount=transaction_amount,
                        created_by=request.user
                    )
                )

                InternalTransferTransaction.objects.create(
                    internal_transfer=internal_transfer,
                    transaction=bank_transaction,
                    role='source'
                )

                InternalTransferTransaction.objects.create(
                    internal_transfer=internal_transfer,
                    transaction=destination_transaction,
                    role='destination'
                )

            created_source_allocations = []
            created_destination_allocations = []

            # -------------------------------------------------
            # Create source Grant allocations
            # -------------------------------------------------

            for item in prepared_allocations:

                grant = grants[item['grant_id']]
                amount = item['amount']

                source_allocation = (
                    TransactionGrantAllocation.objects.create(
                        transaction=bank_transaction,
                        grant=grant,
                        amount=amount,
                        allocation_type='destination',
                        created_by=request.user
                    )
                )

                created_source_allocations.append(
                    source_allocation
                )

                # ---------------------------------------------
                # Internal transfer:
                # assign same Grant to destination credit
                # ---------------------------------------------

                if destination_transaction:

                    destination_allocation = (
                        TransactionGrantAllocation.objects.create(
                            transaction=destination_transaction,
                            grant=grant,
                            amount=amount,
                            allocation_type='source',
                            created_by=request.user
                        )
                    )

                    created_destination_allocations.append(
                        destination_allocation
                    )

        # =====================================================
        # 14. Response
        # =====================================================

        response_data = {

            'success': True,

            'message': (
                'Grants assigned successfully.'
            ),

            'transaction': {
                'id': bank_transaction.id,
                'amount': str(transaction_amount),
                'transaction_type': (
                    bank_transaction.txn_type
                )
            },

            'allocations': [
                {
                    'grant_id': allocation.grant_id,
                    'grant_name': allocation.grant.name,
                    'amount': str(allocation.amount)
                }
                for allocation
                in created_source_allocations
            ]
        }

        # -----------------------------------------------------
        # Internal transfer response
        # -----------------------------------------------------

        if destination_transaction:

            response_data['internal_transfer'] = {

                'id': internal_transfer.id,

                'source_transaction_id': (
                    bank_transaction.id
                ),

                'destination_transaction_id': (
                    destination_transaction.id
                ),

                'amount': str(transaction_amount)
            }

            response_data['destination'] = {

                'transaction_id': (
                    destination_transaction.id
                ),

                'allocations': [
                    {
                        'grant_id': allocation.grant_id,
                        'grant_name': allocation.grant.name,
                        'amount': str(allocation.amount)
                    }
                    for allocation
                    in created_destination_allocations
                ]
            }

        return Response(
            response_data,
            status=status.HTTP_201_CREATED
        )

    except Exception as e:

        logger.error(
            f'Error assigning Grants to debit '
            f'transaction {transaction_id}: {e}',
            exc_info=True
        )

        return Response(
            {
                'success': False,
                'error': str(e)
            },
            status=status.HTTP_400_BAD_REQUEST
        )





@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_bank_grants_funds(request, bank_account_id):

    company = request.user.company

    if not company:
        return Response(
            {
                'success': False,
                'error': 'User is not associated with a company.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # -----------------------------------------
    # Check bank account belongs to company
    # -----------------------------------------

    try:
        bank_account = BankAccount.objects.get(
            id=bank_account_id,
            company=company
        )
    except BankAccount.DoesNotExist:
        return Response(
            {
                'success': False,
                'error': 'Bank account not found.'
            },
            status=status.HTTP_404_NOT_FOUND
        )

    # -----------------------------------------
    # Get funds related to this bank
    # -----------------------------------------

    fund_ids = (
        TransactionFundAllocation.objects
        .filter(
            transaction__bank_account=bank_account,
            fund__grant__company=company
        )
        .values_list(
            'fund_id',
            flat=True
        )
        .distinct()
    )

    funds = (
        Fund.objects
        .filter(
            id__in=fund_ids,
            grant__company=company
        )
        .select_related('grant')
        .order_by(
            'grant__name',
            'name'
        )
    )

    # -----------------------------------------
    # Group Funds under Grants
    # -----------------------------------------

    grants = {}

    for fund in funds:

        grant_id = fund.grant_id

        if grant_id not in grants:
            grants[grant_id] = {
                'grant_id': grant_id,
                'grant_name': fund.grant.name,
                'funds': []
            }

        grants[grant_id]['funds'].append(
            {
                'fund_id': fund.id,
                'fund_name': fund.name
            }
        )

    return Response(
        {
            'success': True,
            'bank_account': {
                'id': bank_account.id,
                'bank_name': bank_account.bank.bank_name,
            },
            'grants': list(grants.values())
        },
        status=status.HTTP_200_OK
    )




def build_grant_outflow_report(self):

    # =========================================================
    # 1. Get debit Grant allocations for selected month
    # =========================================================

    queryset = (
        TransactionGrantAllocation.objects
        .filter(
            transaction__company=self.company,
            transaction__txn_type='debit',
            transaction__txn_date__range=[
                self.start_date,
                self.end_date,
            ],
            allocation_type='destination',
            grant__is_active=True,
        )
        .select_related(
            'transaction',
            'grant',
            'grant__agency',
        )
        .order_by(
            'grant__agency__name',
            'grant__name',
        )
    )

    # ---------------------------------------------------------
    # Optional bank account filter
    # ---------------------------------------------------------

    if self.bank_account:

        queryset = queryset.filter(
            transaction__bank_account_id=self.bank_account
        )

    # =========================================================
    # 2. Group selected-month outflow by Grant
    # =========================================================

    grant_data = {}

    for allocation in queryset:

        grant = allocation.grant
        transaction = allocation.transaction

        if grant.id not in grant_data:

            grant_data[grant.id] = {

                'grant_id': grant.id,

                'grant_name': grant.name,

                'agency_id': (
                    grant.agency.id
                    if grant.agency
                    else None
                ),

                'agency_name': (
                    grant.agency.name
                    if grant.agency
                    else None
                ),

                # ---------------------------------------------
                # Grant's configured amount
                # ---------------------------------------------

                'grant_amount': (
                    grant.amount
                    or Decimal('0.00')
                ),

                # ---------------------------------------------
                # Selected month's outflow
                # ---------------------------------------------

                'monthly_outflow': Decimal('0.00'),

                'weekly_amounts': {
                    f'week_{week["week"]}': Decimal('0.00')
                    for week in self.weeks
                },
            }

        # -----------------------------------------------------
        # Find transaction week
        # -----------------------------------------------------

        for week in self.weeks:

            if (
                week['start']
                <= transaction.txn_date
                <= week['end']
            ):

                week_key = f"week_{week['week']}"

                grant_data[
                    grant.id
                ]['weekly_amounts'][week_key] += (
                    allocation.amount
                )

                grant_data[
                    grant.id
                ]['monthly_outflow'] += (
                    allocation.amount
                )

                break

    # =========================================================
    # 3. Build final result
    # =========================================================

    results = []

    for data in grant_data.values():

        grant_id = data['grant_id']

        # =====================================================
        # Calculate ALL-TIME debit usage
        #
        # IMPORTANT:
        # No month restriction here.
        # =====================================================

        destination_filter = {
            'grant_id': grant_id,

            'allocation_type': 'destination',

            'transaction__company': self.company,

            'transaction__txn_type': 'debit',
        }

        # -----------------------------------------------------
        # If a bank is selected, remaining is calculated
        # only for that bank.
        # -----------------------------------------------------

        if self.bank_account:

            destination_filter[
                'transaction__bank_account_id'
            ] = self.bank_account

        total_used_amount = (
            TransactionGrantAllocation.objects
            .filter(
                **destination_filter
            )
            .aggregate(
                total=Sum('amount')
            )['total']
            or Decimal('0.00')
        )

        total_used_amount = Decimal(
            str(total_used_amount)
        )

        # =====================================================
        # Remaining Grant amount
        #
        # Grant Amount - ALL historical debit usage
        # =====================================================

        remaining_amount = (
            data['grant_amount']
            - total_used_amount
        )

        if remaining_amount < Decimal('0.00'):

            remaining_amount = Decimal('0.00')

        # =====================================================
        # Final response row
        # =====================================================

        item = {

            'grant_id': data['grant_id'],

            'grant_name': data['grant_name'],

            'agency_id': data['agency_id'],

            'agency_name': data['agency_name'],

            # Full configured Grant amount
            'grant_amount': str(
                data['grant_amount']
            ),

            # Selected month's outflow
            'total_outflow': str(
                data['monthly_outflow']
            ),

            # Remaining based on complete history
            'remaining_amount': str(
                remaining_amount
            ),
        }

        # -----------------------------------------------------
        # Weekly values
        # -----------------------------------------------------

        for week in self.weeks:

            key = f"week_{week['week']}"

            item[key] = float(
                data['weekly_amounts'][key]
            )

        results.append(item)

    return results



# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def grant_wise_outflow(request):

#     company = request.user.company

#     # =====================================================
#     # 1. Validate company
#     # =====================================================

#     if not company:

#         return Response(
#             {
#                 'success': False,
#                 'error': (
#                     'User is not associated '
#                     'with a company.'
#                 )
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     # =====================================================
#     # 2. Request parameters
#     # =====================================================

#     month = request.query_params.get(
#         'month'
#     )

#     year = request.query_params.get(
#         'year'
#     )

#     bank_account = request.query_params.get(
#         'bank_account'
#     )

#     grant_id = request.query_params.get(
#         'grant_id'
#     )

#     # =====================================================
#     # 3. Validate Bank Account
#     # =====================================================

#     if bank_account:

#         try:

#             BankAccount.objects.get(
#                 id=bank_account,
#                 company=company
#             )

#         except BankAccount.DoesNotExist:

#             return Response(
#                 {
#                     'success': False,
#                     'error': (
#                         'Bank account not found '
#                         'for this company.'
#                     )
#                 },
#                 status=status.HTTP_404_NOT_FOUND
#             )

#     # =====================================================
#     # 4. Validate Grant
#     # =====================================================

#     if grant_id:

#         try:

#             Grant.objects.get(
#                 id=grant_id,
#                 company=company,
#                 is_active=True
#             )

#         except Grant.DoesNotExist:

#             return Response(
#                 {
#                     'success': False,
#                     'error': (
#                         'Grant not found or inactive.'
#                     )
#                 },
#                 status=status.HTTP_404_NOT_FOUND
#             )

#     # =====================================================
#     # 5. Create Grant-wise service
#     # =====================================================

#     service = GrantWiseOutflowService(
#         company=company,
#         month=month,
#         year=year,
#         bank_account=bank_account,
#         grant_id=grant_id
#     )

#     # =====================================================
#     # 6. Execute
#     # =====================================================

#     try:

#         result = service.execute()

#         return Response(
#             result,
#             status=status.HTTP_200_OK
#         )

#     except ValueError as e:

#         return Response(
#             {
#                 'success': False,
#                 'error': str(e)
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     except Exception as e:

#         logger.error(
#             (
#                 f'Error generating Grant-wise '
#                 f'outflow: {e}'
#             ),
#             exc_info=True
#         )

#         return Response(
#             {
#                 'success': False,
#                 'error': (
#                     'Failed to generate '
#                     'Grant-wise outflow.'
#                 )
#             },
#             status=status.HTTP_500_INTERNAL_SERVER_ERROR
#         )



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def grant_wise_outflow(request):

    # =====================================================
    # 1. Get Company
    # =====================================================

    company = request.user.company

    if not company:

        return Response(
            {
                'success': False,

                'error': (
                    'User is not associated '
                    'with a company.'
                ),
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # =====================================================
    # 2. Get Parameters
    # =====================================================

    month = request.query_params.get(
        'month'
    )

    year = request.query_params.get(
        'year'
    )

    bank_account_id = (
        request.query_params.get(
            'bank_account_id'
        )
    )

    grant_id = (
        request.query_params.get(
            'grant_id'
        )
    )

    # =====================================================
    # 3. Required Parameters
    # =====================================================

    if not month:

        return Response(
            {
                'success': False,

                'error': (
                    'month is required.'
                ),
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    if not year:

        return Response(
            {
                'success': False,

                'error': (
                    'year is required.'
                ),
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    if not bank_account_id:

        return Response(
            {
                'success': False,

                'error': (
                    'bank_account_id is required.'
                ),
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    if not grant_id:

        return Response(
            {
                'success': False,

                'error': (
                    'grant_id is required.'
                ),
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # =====================================================
    # 4. Validate Bank
    # =====================================================

    if not BankAccount.objects.filter(
        id=bank_account_id,
        company=company
    ).exists():

        return Response(
            {
                'success': False,

                'error': (
                    'Bank account not found '
                    'for this company.'
                ),
            },
            status=status.HTTP_404_NOT_FOUND
        )

    # =====================================================
    # 5. Validate Grant
    # =====================================================

    if not Grant.objects.filter(
        id=grant_id,
        company=company,
        is_active=True
    ).exists():

        return Response(
            {
                'success': False,

                'error': (
                    'Grant not found or inactive.'
                ),
            },
            status=status.HTTP_404_NOT_FOUND
        )

    # =====================================================
    # 6. Create Service
    # =====================================================

    service = GrantWiseOutflowService(
        company=company,

        month=month,

        year=year,

        bank_account=bank_account_id,

        grant_id=grant_id,
    )

    # =====================================================
    # 7. Execute
    # =====================================================

    try:

        result = service.execute()

        return Response(
            result,
            status=status.HTTP_200_OK
        )

    except ValueError as e:

        return Response(
            {
                'success': False,

                'error': str(e),
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    except Exception as e:

        logger.error(
            (
                'Error generating Grant-wise '
                f'outflow: {e}'
            ),
            exc_info=True
        )

        return Response(
            {
                'success': False,

                'error': (
                    'Failed to generate '
                    'Grant-wise outflow.'
                ),
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )





@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_grant_wise_outflow(request):

    try:

        # =====================================================
        # 1. Get Company
        # =====================================================

        company = request.user.company

        if not company:

            return Response(
                {
                    'success': False,
                    'error': (
                        'User is not associated '
                        'with a company.'
                    ),
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # =====================================================
        # 2. Get Parameters
        # =====================================================

        month = request.query_params.get(
            'month'
        )

        year = request.query_params.get(
            'year'
        )

        bank_account_id = request.query_params.get(
            'bank_account_id'
        )

        grant_id = request.query_params.get(
            'grant_id'
        )

        # =====================================================
        # 3. Validate Required Parameters
        # =====================================================

        if not month:

            return Response(
                {
                    'success': False,
                    'error': 'month is required.',
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        if not year:

            return Response(
                {
                    'success': False,
                    'error': 'year is required.',
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        if not bank_account_id:

            return Response(
                {
                    'success': False,
                    'error': (
                        'bank_account_id is required.'
                    ),
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        if not grant_id:

            return Response(
                {
                    'success': False,
                    'error': (
                        'grant_id is required.'
                    ),
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # =====================================================
        # 4. Validate Bank Account
        # =====================================================

        if not BankAccount.objects.filter(
            id=bank_account_id,
            company=company
        ).exists():

            return Response(
                {
                    'success': False,
                    'error': (
                        'Bank account not found '
                        'for this company.'
                    ),
                },
                status=status.HTTP_404_NOT_FOUND
            )

        # =====================================================
        # 5. Validate Grant
        # =====================================================

        if not Grant.objects.filter(
            id=grant_id,
            company=company,
            is_active=True
        ).exists():

            return Response(
                {
                    'success': False,
                    'error': (
                        'Grant not found or inactive.'
                    ),
                },
                status=status.HTTP_404_NOT_FOUND
            )

        # =====================================================
        # 6. Create SAME Existing Service
        #
        # No separate calculation logic.
        # =====================================================

        service = GrantWiseOutflowService(

            company=company,

            month=month,

            year=year,

            bank_account=bank_account_id,

            grant_id=grant_id,
        )

        # =====================================================
        # 7. Execute SAME Logic
        # =====================================================

        result = service.execute()

        # =====================================================
        # 8. Extract Weeks
        # =====================================================

        weeks = result.get(
            'weeks',
            []
        )

        # =====================================================
        # 9. Extract Grant Wise Outflow
        # =====================================================

        grant_wise_outflow = result.get(
            'grant_wise_outflow',
            {}
        )

        report_rows = grant_wise_outflow.get(
            'rows',
            []
        )

        totals = grant_wise_outflow.get(
            'totals',
            {}
        )

        # =====================================================
        # 10. Get Particular Rows
        #
        # Your service stores the particulars inside
        # the first report row.
        # =====================================================

        particulars = []

        if report_rows:

            particulars = report_rows[0].get(
                'particulars',
                []
            )

        # =====================================================
        # 11. Create Workbook
        # =====================================================

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = (
            'Grant Wise Outflow'
        )

        # =====================================================
        # 12. Calculate Number of Columns
        #
        # Particular + Weeks + Total
        # =====================================================

        total_columns = (
            1
            + len(weeks)
            + 1
        )

        # =====================================================
        # 13. Report Title
        # =====================================================

        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=total_columns
        )

        worksheet['A1'] = (
            f'Grant Wise Outflow Report '
            f'- {month}/{year}'
        )

        worksheet['A1'].font = Font(
            bold=True,
            size=14
        )

        worksheet['A1'].alignment = Alignment(
            horizontal='center'
        )

        # =====================================================
        # 14. Empty Row
        # =====================================================

        worksheet.append([])

        # =====================================================
        # 15. Table Headers
        # =====================================================

        headers = [

            'Particular'

        ]

        for week in weeks:

            headers.append(
                week.get(
                    'label',
                    (
                        f"Week "
                        f"{week.get('week')}"
                    )
                )
            )

        headers.append(
            'Total'
        )

        worksheet.append(
            headers
        )

        # =====================================================
        # 16. Style Header
        # =====================================================

        header_row = worksheet.max_row

        for cell in worksheet[header_row]:

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal='center',
                vertical='center'
            )

        # =====================================================
        # 17. Add Particular Rows
        # =====================================================

        for particular in particulars:

            excel_row = [

                particular.get(
                    'particular_name',
                    'Uncategorized'
                )

            ]

            # -------------------------------------------------
            # Weekly Amounts
            # -------------------------------------------------

            for week in weeks:

                week_key = (
                    f"week_{week['week']}"
                )

                excel_row.append(

                    particular.get(
                        week_key,
                        0
                    )
                )

            # -------------------------------------------------
            # Monthly Total
            # -------------------------------------------------

            excel_row.append(

                particular.get(
                    'total',
                    0
                )
            )

            worksheet.append(
                excel_row
            )

        # =====================================================
        # 18. Add Total Row
        # =====================================================

        total_row = [

            'Total'

        ]

        # -----------------------------------------------------
        # Weekly Totals
        # -----------------------------------------------------

        for week in weeks:

            week_key = (
                f"week_{week['week']}"
            )

            total_row.append(

                totals.get(
                    week_key,
                    0
                )
            )

        # -----------------------------------------------------
        # Monthly Total
        # -----------------------------------------------------

        total_row.append(

            totals.get(
                'total',
                0
            )
        )

        worksheet.append(
            total_row
        )

        # =====================================================
        # 19. Style Total Row
        # =====================================================

        total_row_number = worksheet.max_row

        for cell in worksheet[
            total_row_number
        ]:

            cell.font = Font(
                bold=True
            )

        # =====================================================
        # 20. Column Width
        # =====================================================

        worksheet.column_dimensions[
            'A'
        ].width = 30

        for column_index in range(
            2,
            total_columns + 1
        ):

            column_letter = (
                worksheet.cell(
                    row=3,
                    column=column_index
                ).column_letter
            )

            worksheet.column_dimensions[
                column_letter
            ].width = 18

        # =====================================================
        # 21. Number Formatting
        # =====================================================

        for row in worksheet.iter_rows(
            min_row=4,
            max_row=worksheet.max_row,
            min_col=2,
            max_col=total_columns
        ):

            for cell in row:

                cell.number_format = (
                    '#,##0.00'
                )

        # =====================================================
        # 22. Save Excel
        # =====================================================

        output = BytesIO()

        workbook.save(
            output
        )

        output.seek(
            0
        )

        # =====================================================
        # 23. Return File
        # =====================================================

        filename = (

            f'grant_wise_outflow_'

            f'{year}_{month}.xlsx'

        )

        response = HttpResponse(

            output.getvalue(),

            content_type=(
                'application/'
                'vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'
            )
        )

        response[
            'Content-Disposition'
        ] = (

            f'attachment; filename="{filename}"'
        )

        return response

    except ValueError as e:

        return Response(
            {
                'success': False,

                'error': str(e),
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    except Exception as e:

        logger.error(

            (
                'Error downloading Grant-wise '
                f'outflow Excel: {e}'
            ),

            exc_info=True
        )

        return Response(
            {
                'success': False,

                'error': (
                    'Failed to download '
                    'Grant-wise outflow report.'
                ),
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )



    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_fund_tracker(request, fund_id):

    company = request.user.company

    if not company:
        return Response(
            {
                'success': False,
                'error': 'User is not associated with a company.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # -----------------------------------------
    # Bank account from query parameter
    # -----------------------------------------

    bank_account_id = request.query_params.get(
        'bank_account_id'
    )

    if not bank_account_id:
        return Response(
            {
                'success': False,
                'error': 'bank_account_id is required.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        bank_account_id = int(bank_account_id)

    except (TypeError, ValueError):
        return Response(
            {
                'success': False,
                'error': 'bank_account_id must be a valid integer.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # -----------------------------------------
    # Bank account
    # -----------------------------------------

    try:
        bank_account = BankAccount.objects.get(
            id=bank_account_id,
            company=company
        )

    except BankAccount.DoesNotExist:
        return Response(
            {
                'success': False,
                'error': 'Bank account not found.'
            },
            status=status.HTTP_404_NOT_FOUND
        )

    # -----------------------------------------
    # Fund
    # -----------------------------------------

    try:
        fund = (
            Fund.objects
            .select_related(
                'grant',
                'source_transaction'
            )
            .get(
                id=fund_id,
                grant__company=company
            )
        )

    except Fund.DoesNotExist:
        return Response(
            {
                'success': False,
                'error': 'Fund not found.'
            },
            status=status.HTTP_404_NOT_FOUND
        )

    # -----------------------------------------
    # Fund allocations for selected bank only
    # -----------------------------------------

    allocations = (
        TransactionFundAllocation.objects
        .filter(
            fund=fund,
            transaction__bank_account=bank_account
        )
        .select_related(
            'transaction',
            'transaction__bank_account',
            'transaction__particular'
        )
        .order_by(
            'transaction__txn_date',
            'transaction__id'
        )
    )

    # -----------------------------------------
    # Amounts
    # -----------------------------------------

    initial_amount = Decimal('0.00')
    transferred_in = Decimal('0.00')
    transferred_out = Decimal('0.00')
    spent_amount = Decimal('0.00')

    transaction_data = []

    for allocation in allocations:

        txn = allocation.transaction
        amount = allocation.amount

        # -------------------------------------
        # Original fund creation
        # -------------------------------------

        if txn.id == fund.source_transaction_id:

            initial_amount += amount

        # -------------------------------------
        # Internal transfer
        # -------------------------------------

        elif txn.matched_txn_id:

            if allocation.allocation_type == 'source':
                transferred_out += amount

            elif allocation.allocation_type == 'destination':
                transferred_in += amount

        # -------------------------------------
        # Normal fund spending
        # -------------------------------------

        elif allocation.allocation_type == 'source':

            spent_amount += amount

        # -------------------------------------
        # Particular
        # -------------------------------------

        particular_id = None
        particular_name = None

        if txn.particular:
            particular_id = txn.particular.id
            particular_name = txn.particular.name

        # -------------------------------------
        # Transaction history
        # -------------------------------------

        transaction_data.append(
            {
                'transaction_id': txn.id,

                'transaction_date': (
                    txn.txn_date.strftime('%d-%m-%Y')
                    if txn.txn_date
                    else None
                ),

                'transaction_type': txn.txn_type,

                'amount': str(amount),

                'allocation_type': allocation.allocation_type,

                'reference_no': txn.ref_no,

                'particular_id': particular_id,

                'particular_name': particular_name,

                'matched_transaction_id': (
                    txn.matched_txn_id
                )
            }
        )

    # -----------------------------------------
    # Remaining balance
    # -----------------------------------------

    remaining_amount = (
        initial_amount
        + transferred_in
        - transferred_out
        - spent_amount
    )

    # -----------------------------------------
    # Response
    # -----------------------------------------

    return Response(
        {
            'success': True,

            'fund': {
                'id': fund.id,
                'name': fund.name,
                'grant_id': fund.grant.id,
                'grant_name': fund.grant.name
            },

            'bank_account': {
                'id': bank_account.id,
                'bank_name': bank_account.bank.bank_name
            },

            'summary': {
                'initial_amount': str(initial_amount),
                'transferred_in': str(transferred_in),
                'transferred_out': str(transferred_out),
                'spent_amount': str(spent_amount),
                'remaining_amount': str(remaining_amount)
            },

            'transactions': transaction_data
        },
        status=status.HTTP_200_OK
    )






# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def get_all_bank_grants_funds(request):

#     company = request.user.company

#     if not company:
#         return Response(
#             {
#                 'success': False,
#                 'error': 'User is not associated with a company.'
#             },
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     bank_accounts = (
#         BankAccount.objects
#         .filter(company=company)
#         .order_by('id')
#     )

#     response_data = []

#     for bank_account in bank_accounts:

#         # -------------------------------------------------
#         # Get funds which have allocations in this bank
#         # -------------------------------------------------

#         allocations = (
#             TransactionFundAllocation.objects
#             .filter(
#                 transaction__bank_account=bank_account,
#                 fund__grant__company=company
#             )
#             .select_related(
#                 'fund',
#                 'fund__grant'
#             )
#             .order_by(
#                 'fund__grant__name',
#                 'fund__name'
#             )
#         )

#         # -------------------------------------------------
#         # Group by Grant -> Fund
#         # -------------------------------------------------

#         grants = {}

#         for allocation in allocations:

#             fund = allocation.fund
#             grant = fund.grant

#             if grant.id not in grants:

#                 grants[grant.id] = {
#                     'grant_id': grant.id,
#                     'grant_name': grant.name,
#                     'funds': []
#                 }

#             # ---------------------------------------------
#             # Prevent duplicate fund entries
#             # ---------------------------------------------

#             existing_fund_ids = {
#                 item['fund_id']
#                 for item in grants[grant.id]['funds']
#             }

#             if fund.id not in existing_fund_ids:

#                 grants[grant.id]['funds'].append(
#                     {
#                         'fund_id': fund.id,
#                         'fund_name': fund.name
#                     }
#                 )

#         # -------------------------------------------------
#         # Add bank only if it has related grants
#         # -------------------------------------------------

#         if grants:

#             response_data.append(
#                 {
#                     'bank_account_id': bank_account.id,
#                     'bank_name': bank_account.bank.bank_name,
#                     'grants': list(grants.values())
#                 }
#             )

#     return Response(
#         {
#             'success': True,
#             'banks': response_data
#         },
#         status=status.HTTP_200_OK
#     )



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_all_bank_grants(request):

    company = request.user.company

    # =====================================================
    # 1. Validate company
    # =====================================================

    if not company:
        return Response(
            {
                'success': False,
                'error': 'User is not associated with a company.'
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # =====================================================
    # 2. Get company bank accounts
    # =====================================================

    bank_accounts = (
        BankAccount.objects
        .filter(
            company=company
        )
        .select_related('bank')
        .order_by('id')
    )

    response_data = []

    # Keep track of Grants already assigned to a bank
    assigned_grant_ids = set()

    # =====================================================
    # 3. Get bank-related Grants
    # =====================================================

    for bank_account in bank_accounts:

        allocations = (
            TransactionGrantAllocation.objects
            .filter(
                transaction__bank_account=bank_account,
                grant__company=company,
                grant__is_active=True
            )
            .select_related(
                'grant',
                'grant__agency'
            )
            .order_by(
                'grant__agency__name',
                'grant__name'
            )
        )

        agencies = {}

        for allocation in allocations:

            grant = allocation.grant
            agency = grant.agency

            if not agency:
                continue

            assigned_grant_ids.add(grant.id)

            # ---------------------------------------------
            # Create Agency
            # ---------------------------------------------

            if agency.id not in agencies:

                agencies[agency.id] = {
                    'agency_id': agency.id,
                    'agency_name': agency.name,
                    'grants': []
                }

            # ---------------------------------------------
            # Prevent duplicate Grant
            # ---------------------------------------------

            existing_grant_ids = {
                item['grant_id']
                for item in agencies[agency.id]['grants']
            }

            if grant.id not in existing_grant_ids:

                agencies[agency.id]['grants'].append(
                    {
                        'grant_id': grant.id,
                        'grant_name': grant.name
                    }
                )

        # ---------------------------------------------
        # Add bank only if it has related Grants
        # ---------------------------------------------

        if agencies:

            response_data.append(
                {
                    'bank_account_id': bank_account.id,

                    'bank_name': (
                        bank_account.bank.bank_name
                        if bank_account.bank
                        else None
                    ),

                    'agencies': list(
                        agencies.values()
                    )
                }
            )

    # =====================================================
    # 4. Get company Grants that are NOT assigned
    #    to any bank
    # =====================================================

    unassigned_grants = (
        Grant.objects
        .filter(
            company=company,
            is_active=True
        )
        .exclude(
            id__in=assigned_grant_ids
        )
        .select_related('agency')
        .order_by(
            'agency__name',
            'name'
        )
    )

    # =====================================================
    # 5. Group unassigned Grants by Agency
    # =====================================================

    unassigned_agencies = {}

    for grant in unassigned_grants:

        agency = grant.agency

        if not agency:
            continue

        if agency.id not in unassigned_agencies:

            unassigned_agencies[agency.id] = {
                'agency_id': agency.id,
                'agency_name': agency.name,
                'grants': []
            }

        unassigned_agencies[agency.id]['grants'].append(
            {
                'grant_id': grant.id,
                'grant_name': grant.name
            }
        )

    # =====================================================
    # 6. Final response
    # =====================================================

    return Response(
        {
            'success': True,

            'banks': response_data,

            'unassigned_grants': list(
                unassigned_agencies.values()
            )
        },
        status=status.HTTP_200_OK
    )






@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def upload_transaction_receipt(request, transaction_id):
    try:
        # -------------------------------------------------
        # 1. Get transaction
        # -------------------------------------------------
        transaction = Transaction.objects.filter(
            id=transaction_id,
            company=request.user.company
        ).first()

        if not transaction:
            return Response(
                {
                    'success': False,
                    'error': 'Transaction not found.'
                },
                status=status.HTTP_404_NOT_FOUND
            )

        # -------------------------------------------------
        # 2. Only debit transactions can have receipts
        # -------------------------------------------------
        if transaction.txn_type != 'debit':
            return Response(
                {
                    'success': False,
                    'error': 'Receipts can only be uploaded for debit transactions.'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # -------------------------------------------------
        # 3. Get uploaded file
        # -------------------------------------------------
        uploaded_file = request.FILES.get('file')

        if not uploaded_file:
            return Response(
                {
                    'success': False,
                    'error': 'file is required'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # -------------------------------------------------
        # 4. Validate file type
        # -------------------------------------------------
        allowed = {
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/pdf',
            'image/jpeg',
            'image/png',
        }

        if uploaded_file.content_type not in allowed:
            return Response(
                {
                    'success': False,
                    'error': f'Unsupported file type: {uploaded_file.content_type}'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # -------------------------------------------------
        # 5. Get company folder
        # -------------------------------------------------
        company_name = _get_company_name(request)

        if not company_name:
            return Response(
                {
                    'success': False,
                    'error': 'Could not determine company name'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        company_result = drive_service.get_or_create_company_folder(
            company_name
        )

        if not company_result:
            return Response(
                {
                    'success': False,
                    'error': 'Could not access company Drive folder'
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # -------------------------------------------------
        # 6. Get Billing_Receipt folder
        # -------------------------------------------------
        billing_folder_id = company_result['subfolders'].get(
            'Billing_Receipt'
        )

        if not billing_folder_id:
            return Response(
                {
                    'success': False,
                    'error': 'Billing_Receipt folder not found'
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # -------------------------------------------------
        # 7. Upload to Drive
        # -------------------------------------------------
        file_bytes = uploaded_file.read()

        result = drive_service.upload_file(
            file_obj=io.BytesIO(file_bytes),
            file_name=uploaded_file.name,
            folder_id=billing_folder_id,
            mime_type=uploaded_file.content_type,
        )

        if not result:
            return Response(
                {
                    'success': False,
                    'error': 'Failed to upload to Drive'
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # -------------------------------------------------
        # 8. Create receipt linked directly to transaction
        # -------------------------------------------------
        receipt = ReceiptDocument.objects.create(
            company=request.user.company,
            matched_transaction=transaction,
            drive_file_id=result.get('file_id'),
            file_name=uploaded_file.name,
            file_link=result.get('file_link'),
            mime_type=uploaded_file.content_type,
            uploaded_at=timezone.now(),
            extracted=False,
        )

        # -------------------------------------------------
        # 9. Process receipt asynchronously
        # -------------------------------------------------
        task = process_receipt_task.delay(receipt.id)

        return Response(
            {
                'success': True,
                'message': 'Receipt uploaded and linked to transaction.',
                'transaction_id': transaction.id,
                'receipt_id': receipt.id,
                'task_id': task.id,
                'status_url': f'/api/tasks/{task.id}/status/',
            },
            status=status.HTTP_202_ACCEPTED
        )

    except Exception as e:
        logger.error(
            f'Error uploading transaction receipt: {e}',
            exc_info=True
        )

        return Response(
            {
                'success': False,
                'error': str(e)
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

    



from decimal import Decimal

from rest_framework.decorators import (
    api_view,
    permission_classes,
)
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from .models import Transaction


from decimal import Decimal

from rest_framework.decorators import (
    api_view,
    permission_classes,
)

from rest_framework.permissions import IsAuthenticated

from rest_framework.response import Response

from rest_framework import status


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_transaction_movement_info(
    request,
    transaction_id
):

    company = request.user.company

    # =====================================================
    # 1. Validate company
    # =====================================================

    if not company:

        return Response(
            {
                'success': False,
                'error': (
                    'User is not associated with a company.'
                )
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # =====================================================
    # 2. Get bank_account_id
    #
    # Example:
    #
    # ?bank_account_id=24
    #
    # This determines which bank side of the
    # transaction the user wants to inspect.
    # =====================================================

    bank_account_id = request.GET.get(
        'bank_account_id'
    )

    # =====================================================
    # 3. Get transaction
    # =====================================================

    try:

        bank_transaction = (
            Transaction.objects
            .select_related(
                'bank_account',
                'bank_account__bank',

                'matched_txn',
                'matched_txn__bank_account',
                'matched_txn__bank_account__bank',
            )
            .get(
                id=transaction_id,
                company=company
            )
        )

    except Transaction.DoesNotExist:

        return Response(
            {
                'success': False,
                'error': 'Transaction not found.'
            },
            status=status.HTTP_404_NOT_FOUND
        )

    # =====================================================
    # 4. Helper: Bank details
    # =====================================================

    def get_bank_details(bank_account):

        if not bank_account:
            return None

        return {

            'account_name': (
                bank_account.account_number
                if bank_account.account_number
                else None
            ),

            'bank_name': (
                bank_account.bank.bank_name
                if bank_account.bank
                else None
            ),

            'bank_id': bank_account.id,
        }

    # =====================================================
    # 5. Internal transfer validation
    # =====================================================

    def is_internal_transfer(transaction):

        return (

            transaction.reconcile_status == 'ignored'

            and

            transaction.matched_txn is not None
        )

    # =====================================================
    # 6. Select correct transaction based on bank
    #
    # IMPORTANT:
    #
    # We use the Transaction.balance from the transaction
    # belonging to the bank selected by the user.
    #
    # We do NOT get the latest bank balance.
    # =====================================================

    selected_transaction = bank_transaction

    if bank_account_id:

        try:

            bank_account_id = int(
                bank_account_id
            )

        except ValueError:

            return Response(
                {
                    'success': False,
                    'error': (
                        'Invalid bank_account_id.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # -------------------------------------------------
        # Current transaction belongs to selected bank
        # -------------------------------------------------

        if (
            bank_transaction.bank_account_id
            == bank_account_id
        ):

            selected_transaction = (
                bank_transaction
            )

        # -------------------------------------------------
        # Matched transaction belongs to selected bank
        # -------------------------------------------------

        elif (

            bank_transaction.matched_txn

            and

            bank_transaction
            .matched_txn
            .bank_account_id
            == bank_account_id
        ):

            selected_transaction = (
                bank_transaction.matched_txn
            )

        # -------------------------------------------------
        # Selected bank is not part of this transaction
        # -------------------------------------------------

        else:

            return Response(
                {
                    'success': False,
                    'error': (
                        'Selected bank account is not '
                        'associated with this transaction.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

    # =====================================================
    # 7. Determine Source and Destination
    # =====================================================

    source = None
    destination = None

    # =====================================================
    # NORMAL TRANSACTION
    # =====================================================

    if not is_internal_transfer(
        selected_transaction
    ):

        # -------------------------------------------------
        # CREDIT / CASH INFLOW
        #
        # Description → Bank
        #
        # Source = Description
        # Destination = Current Bank
        # -------------------------------------------------

        if selected_transaction.txn_type == 'credit':

            source = (
                selected_transaction.description
                if selected_transaction.description
                else None
            )

            destination = get_bank_details(
                selected_transaction.bank_account
            )

        # -------------------------------------------------
        # DEBIT / CASH OUTFLOW
        #
        # Bank → Description
        #
        # Source = Current Bank
        # Destination = Description
        # -------------------------------------------------

        elif selected_transaction.txn_type == 'debit':

            source = get_bank_details(
                selected_transaction.bank_account
            )

            destination = (
                selected_transaction.description
                if selected_transaction.description
                else None
            )

    # =====================================================
    # INTERNAL TRANSFER
    #
    # Bank A → Bank B
    #
    # Debit side:
    #
    # Source = Current Bank
    # Destination = Matched Bank
    #
    # Credit side:
    #
    # Source = Matched Bank
    # Destination = Current Bank
    # =====================================================

    else:

        # -------------------------------------------------
        # DEBIT SIDE
        # -------------------------------------------------

        if selected_transaction.txn_type == 'debit':

            source = get_bank_details(
                selected_transaction.bank_account
            )

            destination = get_bank_details(

                selected_transaction
                .matched_txn
                .bank_account

                if selected_transaction.matched_txn

                else None
            )

        # -------------------------------------------------
        # CREDIT SIDE
        # -------------------------------------------------

        elif selected_transaction.txn_type == 'credit':

            source = get_bank_details(

                selected_transaction
                .matched_txn
                .bank_account

                if selected_transaction.matched_txn

                else None
            )

            destination = get_bank_details(
                selected_transaction.bank_account
            )

    # =====================================================
    # 8. Transaction Status
    # =====================================================

    if is_internal_transfer(
        selected_transaction
    ):

        transaction_status = (
            'Internal Transfer'
        )

    elif (
        selected_transaction.reconcile_status
        == 'matched'
    ):

        transaction_status = (
            'Matched'
        )

    elif (
        selected_transaction.reconcile_status
        == 'receipt_missing'
    ):

        transaction_status = (
            'Receipt Missing'
        )

    else:

        transaction_status = (
            'Unmatched'
        )

    # =====================================================
    # 9. Transaction Type
    #
    # IMPORTANT:
    #
    # Do NOT change transaction type to
    # "Internal Transfer".
    #
    # Keep original transaction type.
    # =====================================================

    if selected_transaction.txn_type == 'credit':

        transaction_type = (
            'Cash Inflow'
        )

    elif selected_transaction.txn_type == 'debit':

        transaction_type = (
            'Cash Outflow'
        )

    else:

        transaction_type = (
            selected_transaction.txn_type
        )

    # =====================================================
    # 10. Response
    # =====================================================

    return Response(
        {
            'success': True,

            'transaction': {

                # -----------------------------------------
                # Selected transaction ID
                # -----------------------------------------

                'id': selected_transaction.id,

                # -----------------------------------------
                # Date
                # -----------------------------------------

                'date': (

                    selected_transaction
                    .txn_date
                    .strftime('%d-%m-%Y')

                    if selected_transaction.txn_date

                    else None
                ),

                # -----------------------------------------
                # Source
                #
                # Credit:
                # Description
                #
                # Debit:
                # Bank object
                # -----------------------------------------

                'source': source,

                # -----------------------------------------
                # Destination
                #
                # Credit:
                # Bank object
                #
                # Debit:
                # Description
                # -----------------------------------------

                'destination': destination,

                # -----------------------------------------
                # Transaction Type
                # -----------------------------------------

                'transaction_type': (
                    transaction_type
                ),

                # -----------------------------------------
                # Amount
                # -----------------------------------------

                'amount': str(

                    selected_transaction.amount

                    if selected_transaction.amount
                    is not None

                    else Decimal('0.00')
                ),

                # -----------------------------------------
                # Historical Closing Balance
                #
                # From the selected transaction record.
                #
                # NOT latest bank balance.
                # -----------------------------------------

                'closing_balance': (

                    str(
                        selected_transaction.balance
                    )

                    if selected_transaction.balance
                    is not None

                    else None
                ),

                # -----------------------------------------
                # Reference Number
                # -----------------------------------------

                'reference_no': (

                    selected_transaction.ref_no

                    if selected_transaction.ref_no

                    else None
                ),

                # -----------------------------------------
                # UTR Number
                # -----------------------------------------

                'utr_no': (

                    selected_transaction.utr_no

                    if selected_transaction.utr_no

                    else None
                ),

                # -----------------------------------------
                # Status
                # -----------------------------------------

                'status': transaction_status,
            }
        },
        status=status.HTTP_200_OK
    )